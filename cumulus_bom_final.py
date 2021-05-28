import tkinter as tk
from tkinter import messagebox as MB
from tkinter import scrolledtext
from tkinter import scrolledtext as sc
import tkinter.ttk as ttk1
import math
import base64
# import netmiko
import time
import threading
import xlsxwriter
import xlrd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border, Alignment, Side
from openpyxl.styles import colors
from openpyxl.cell import Cell
from openpyxl.styles import Font
from openpyxl.styles import colors
from openpyxl.cell import Cell
import pandas
import xlwt
from xlutils.copy import copy
from datetime import datetime
import pyexcel
# for mail
import email, smtplib, ssl
import cisco
import cumulus_bom
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from PIL import ImageTk, Image
D={}


def NGC_CUMULUS_Function(root,font,country_arg,amex_id_arg,user_arg,building_name_arg,flr_txt_arg, New_NGC_page_frame):
    countryVar = country_arg.get()
    amex_idvar = amex_id_arg.get()
    usertext = user_arg.get()
    buildind = building_name_arg.get()
    print("flr_txt_arg.get()",flr_txt_arg.get())
    flr = int(flr_txt_arg.get())
    per_floor_usr_count_list = []
    floor_name_list = []
    floor_name_list2 = []
    NGC_frame_for_floor_input = tk.Frame(root, bd=6, relief=tk.SUNKEN, bg='black')
    NGC_frame_for_floor_input.place(x=0.1, y=0.1, relx=0.5, rely=0.4, anchor='n')
    floor1_txt_frame = tk.Frame(root, bd=6, relief=tk.SUNKEN, bg='black',fg='White')
    floor1_txt_frame.place(relx=0.5, rely=0.2, anchor='n')
    floor2_txt_frame = tk.Frame(root, bd=6, relief=tk.SUNKEN, bg='black',fg='White')
    floor2_txt_frame.place(relx=0.5, rely=0.8, anchor='n')
    for i in range(1, flr + 1):
        lable = tk.Label(NGC_frame_for_floor_input, text="Enter Floor Number", font=font,bg='black',fg='White')
        lable.grid(row=10 + i, column=1)
        flr_num_txt = tk.Entry(NGC_frame_for_floor_input, width=30)
        flr_num_txt.grid(row=10 + i, column=2)
        floor_name_list.append(flr_num_txt)
        l = tk.Label(NGC_frame_for_floor_input, text="Enter the EUC count ".format(i), font=font, bg='black',fg='White')
        l.grid(row=10 + i, column=5)
        per_flr_usr_count_txt = tk.Entry(NGC_frame_for_floor_input, width=30)
        per_flr_usr_count_txt.grid(row=10 + i, column=6)
        per_floor_usr_count_list.append(per_flr_usr_count_txt)

        def NGC_CUMULUS_Leaf_Calculation():
            count = False
            for i in range(len(floor_name_list)):
                floor_name_list2.append(floor_name_list[i].get())
                print(floor_name_list2)
            for i1 in floor_name_list2:
                if floor_name_list2.count(i1) >= 2:
                    count = True
                    print("Count:", count)
                    break
            if count == True:
                res = MB.askretrycancel('Message title', 'Floor name same')
                if res == True:
                    print('check')
                    NGC_CUMULUS_Function()
            else:
                # -------------------------------------------------
                #       CODE FOR EXCEL FILE
                # -------------------------------------------------

                def XL_CUMULUS_Small_Site_Type():
                    start_row = 0
                    end_row12 = 1
                    # -------------------------------------------------
                    #       CODE FOR creating user access layer and core layer Dictionaries
                    # -------------------------------------------------
                    wb = xlrd.open_workbook('BOM Code.xls')
                    sh = wb.sheet_by_name('DETAIL-CUMULUS-BOM-SMALL-SITE')
                    num_rows = sh.nrows
                    sh1 = wb.sheet_by_name('SUMMARY-CUMULUS-BOM')
                    num_rows1 = sh1.nrows

                    num_cells = sh.ncols - 1
                    num_cells1 = sh1.ncols - 1
                    curr_row = 0
                    # -------------------------------------------------
                    #       CODE FOR writing back to xl
                    # -------------------------------------------------
                    print("xl created")
                    rb = xlrd.open_workbook("BOM Code.xls")
                    sh = rb.sheet_by_name('DETAIL-CUMULUS-BOM-SMALL-SITE')
                    sh1 = rb.sheet_by_name('SUMMARY-CUMULUS-BOM')
                    wb = copy(rb)
                    w_sheet = wb.get_sheet('DETAIL-CUMULUS-BOM-SMALL-SITE')
                    w_sheet1 = wb.get_sheet('SUMMARY-CUMULUS-BOM')

                    # -------------------------------------------------
                    #       CODE TO GET VALUES INTO CUMULUS-SUMMARY-TAB
                    # -------------------------------------------------
                    # SUMMARY-PAGE
                    for i in range(2, num_rows1):
                        partnumber = sh1.cell(i, 2).value
                        if partnumber == '407-BBOU-US':
                            if flr == 1:
                                v = leaf_list1 * 4 + 22 + 2
                                w_sheet1.write(i, 4, v)
                            else:
                                v = leaf_list1 * 4 + 22 + flr * 4 + 6
                                w_sheet1.write(i, 4, v)

                        if partnumber == '407-BBOS-US':
                            v = 34
                            w_sheet1.write(i, 4, v)
                        if partnumber == '210-ADUX':
                            if flr == 1:
                                v = 2 + 1
                                w_sheet1.write(i, 4, v)
                            else:
                                v = 2 + 1 + 2
                                w_sheet1.write(i, 4, v)

                        if partnumber == '4610-54P-O-AC-F-US':
                            v = leaf_list1 + 1 + flr
                            w_sheet1.write(i, 4, v)

                        if partnumber == 'A9748229':
                            if flr == 1:
                                v = 2
                                w_sheet1.write(i, 4, v)
                            else:
                                v = 4
                                w_sheet1.write(i, 4, v)

                        if partnumber == 'A8793201':
                            v = leaf_list1 + flr
                            w_sheet1.write(i, 4, v)

                        if partnumber == 'SLC80482201S':
                            if flr == 1:
                                v = flr
                                w_sheet1.write(i, 4, v)
                            else:
                                v = flr + 2
                                w_sheet1.write(i, 4, v)

                    # -------------------------------------------------
                    #       CODE TO GET VALUES OF ROW and insert leaf value inside qty column
                    # -------------------------------------------------
                    # CORE LAYER ENTRY
                    for i in range(2, num_rows):
                        partnumber = sh.cell(i, 0).value
                        if partnumber == "":
                            end_row = i
                            break
                    core_layer_end_row = end_row

                    print("end", end_row)

                    for i in range(2, core_layer_end_row):
                        partnumber = sh.cell(i, 2).value
                        if partnumber == '407-BBOU-US':
                            v = leaf_list1 * 2 + 22
                            w_sheet.write(i, 4, v)
                        if partnumber == '407-BBOS-US':
                            v = 30
                            w_sheet.write(i, 4, v)
                        if partnumber == '210-ADUX':
                            v = 2 + 1
                            w_sheet.write(i, 4, v)
                        if partnumber == 'A9748229':
                            v = 2
                            w_sheet.write(i, 4, v)

                    # USER ACCESS LAYER ENTRY
                    for i in range(2, num_rows):
                        for j in range(0, num_cells):
                            partnumber = sh.cell(i, j).value
                            if partnumber == 'User Access Leaves':
                                start_row = i
                                break
                    start_row1 = start_row

                    for i in range(start_row1, num_rows):
                        partnumber = sh.cell(i, 0).value
                        if partnumber == "":
                            end_row12 = i
                            break
                    USR_Access_end_row = end_row12

                    for i in range(start_row1, USR_Access_end_row):
                        partnumber = sh.cell(i, 2).value

                        if partnumber == '407-BBOU-US':
                            v = leaf_list1 * 2
                            w_sheet.write(i, 4, v)

                        if partnumber == '4610-54P-O-AC-F-US':
                            v = leaf_list1 + 1
                            w_sheet.write(i, 4, v)
                        if partnumber == 'A8793201':
                            v = leaf_list1
                            w_sheet.write(i, 4, v)

                    # MGMT CORE  LAYER ENTRY

                    if flr == 1:
                        for i in range(2, num_rows):
                            for j in range(0, num_cells):
                                partnumber = sh.cell(i, j).value
                                if partnumber == 'Management Core Layer':
                                    start_row = i
                                    break
                        start_row1 = start_row

                        for i in range(start_row1, num_rows):
                            partnumber = sh.cell(i, 0).value
                            if partnumber == "":
                                end_row12 = i
                                break
                        OOB_CORE_end_row = end_row12
                        for i in range(start_row1, OOB_CORE_end_row):
                            partnumber = sh.cell(i, 2).value
                            if partnumber == '407-BBOU-US':
                                if flr == 1:
                                    v = flr * 2
                                    w_sheet.write(i, 4, v)
                                else:
                                    v = flr * 2 + 6
                                    w_sheet.write(i, 4, v)

                            if partnumber == '407-BBOS-US':
                                v = 4
                                w_sheet.write(i, 4, v)
                            if partnumber == '210-ADUX':
                                v = 1
                                w_sheet.write(i, 4, v)

                                b = 'EDGECORE NETWORKS'
                                w_sheet.write(i, 1, b)
                                c = '4610-54P-O-AC-F-US'
                                w_sheet.write(i, 2, c)
                                d = 'AS4610-54P 48 PORT 10/100/1000BASE-T SWITCH WITH 48 POE+ PORTS,1-8 PORT SUPPORT'
                                w_sheet.write(i, 3, d)

                            if partnumber == 'A9748229':
                                v = 1
                                w_sheet.write(i, 4, v)
                                c1 = 'A8793201'
                                w_sheet.write(i, 2, c1)
                                d1 = 'Cumulus Linux Perpetual License, 1G, 1 Year Software Updates and Support Included'
                                w_sheet.write(i, 3, d1)

                        # Terminal Server Entry
                        for i in range(2, num_rows):
                            for j in range(0, num_cells):
                                partnumber = sh.cell(i, j).value
                                if partnumber == 'Terminal Server':
                                    start_row = i
                                    break
                        start_row1 = start_row

                        for i in range(start_row1, num_rows):
                            partnumber = sh.cell(i, 0).value
                            if partnumber == "":
                                end_row12 = i
                                break
                        OOB_CORE_end_row = end_row12
                        for i in range(start_row1, OOB_CORE_end_row):
                            partnumber = sh.cell(i, 2).value
                            if partnumber == 'SLC80482201S':
                                v = flr
                                w_sheet.write(i, 4, v)

                        # MGMT Access LAYER REMOVAL

                        for i in range(2, num_rows):
                            for j in range(0, num_cells):
                                partnumber = sh.cell(i, j).value
                                if partnumber == 'Management Access Leaves':
                                    start_row = i
                                    break
                        start_row1 = start_row

                        for i in range(OOB_CORE_end_row, num_rows):
                            for j in range(num_cells + 1):
                                w_sheet.write(i, j, "")

                    else:
                        for i in range(2, num_rows):
                            for j in range(0, num_cells):
                                partnumber = sh.cell(i, j).value
                                if partnumber == 'Management Core Layer':
                                    start_row = i
                                    break
                        start_row1 = start_row

                        for i in range(start_row1, num_rows):
                            partnumber = sh.cell(i, 0).value
                            if partnumber == "":
                                end_row12 = i
                                break
                        OOB_CORE_end_row = end_row12
                        for i in range(start_row1, OOB_CORE_end_row):
                            partnumber = sh.cell(i, 2).value
                            if partnumber == '407-BBOU-US':
                                v = flr * 2 + 6
                                w_sheet.write(i, 4, v)
                            if partnumber == '407-BBOS-US':
                                v = 4
                                w_sheet.write(i, 4, v)
                            if partnumber == '210-ADUX':
                                v = 2
                                w_sheet.write(i, 4, v)

                            if partnumber == 'A9748229':
                                v = 2
                                w_sheet.write(i, 4, v)

                        # Terminal Server Entry
                        for i in range(2, num_rows):
                            for j in range(0, num_cells):
                                partnumber = sh.cell(i, j).value
                                if partnumber == 'Terminal Server':
                                    start_row = i
                                    break
                        start_row1 = start_row

                        for i in range(start_row1, num_rows):
                            partnumber = sh.cell(i, 0).value
                            if partnumber == "":
                                end_row12 = i
                                break
                        OOB_CORE_end_row = end_row12
                        for i in range(start_row1, OOB_CORE_end_row):
                            partnumber = sh.cell(i, 2).value
                            if partnumber == 'SLC80482201S':
                                v = flr + 2
                                w_sheet.write(i, 4, v)

                        # MGMT Leaves LAYER ENTRY
                        for i in range(2, num_rows):
                            for j in range(0, num_cells):
                                partnumber = sh.cell(i, j).value
                                if partnumber == 'Management Access Leaves':
                                    start_row = i
                                    break
                        start_row1 = start_row

                        for i in range(start_row1, num_rows):

                            partnumber = sh.cell(i, 0).value
                            if partnumber == "":
                                end_row12 = i
                                break
                        OOB_Leaves_end_row = end_row12
                        for i in range(start_row1, num_rows):
                            partnumber = sh.cell(i, 2).value
                            if partnumber == '407-BBOU-US':
                                v = flr * 2
                                w_sheet.write(i, 4, v)

                            if partnumber == '4610-54P-O-AC-F-US' or partnumber == 'A8793201':
                                v = flr
                                w_sheet.write(i, 4, v)

                    # -------------------------------------------------
                    #       CODE FOR saving xl file after updating
                    # -------------------------------------------------
                    now = datetime.now()
                    s = str('result{}.xls'.format(now.strftime("%c")))
                    wb.save(s.replace(':', '_'))
                    pyexcel.save_book_as(file_name=s.replace(':', '_'),
                                         dest_file_name='result{}.xlsx'.format(now.strftime("%c")).replace(':',
                                                                                                           '_'))
                    s = 'result{}.xlsx'.format(now.strftime("%c")).replace(':', '_')

                    # -------------------------------------------------
                    #       Used openpyxl for editing the xlsx file
                    # -------------------------------------------------

                    import openpyxl
                    from openpyxl import Workbook

                    wb = openpyxl.load_workbook(s)
                    sheet = wb['DETAIL-CUMULUS-BOM-SMALL-SITE']
                    sheet1 = wb['SUMMARY-CUMULUS-BOM']

                    # -------------------------------------------------
                    #       CODE FOR col width(adjust the col width according to text length in it)
                    # -------------------------------------------------
                    dims = {}
                    for row in sheet.rows:

                        for cell in row:
                            if cell.value:
                                dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))
                    for col, value in dims.items():
                        sheet.column_dimensions[col].width = value

                    sheet.column_dimensions['E'].width = 10
                    sheet1.column_dimensions['A'].width = 15
                    sheet1.column_dimensions['B'].width = 25
                    sheet1.column_dimensions['C'].width = 25
                    sheet1.column_dimensions['D'].width = 110
                    sheet1.column_dimensions['E'].width = 10

                    # -------------------------------------------------
                    #      Cell Formatting for small site
                    # -------------------------------------------------

                    top_left11_cell = sheet1['A2']

                    top_left_cell = sheet['A2']
                    top_left1_cell = sheet['A9']
                    top_left2_cell = sheet['A15']
                    top_left3_cell = sheet['A22']
                    top_left4_cell = sheet['A26']
                    sheet.merge_cells('A2:E2')
                    sheet.merge_cells('A9:E9')
                    sheet.merge_cells('A15:E15')
                    sheet.merge_cells('A22:E22')
                    sheet.merge_cells('A26:E26')

                    sheet1.merge_cells('A2:E2')

                    top_left_cell.alignment = Alignment(horizontal="center", vertical="center")
                    top_left1_cell.alignment = Alignment(horizontal="center", vertical="center")
                    top_left2_cell.alignment = Alignment(horizontal="center", vertical="center")
                    top_left3_cell.alignment = Alignment(horizontal="center", vertical="center")
                    top_left4_cell.alignment = Alignment(horizontal="center", vertical="center")

                    top_left11_cell.alignment = Alignment(horizontal="center", vertical="center")

                    sheet.cell(row=2, column=1).fill = PatternFill(start_color='FFFF99', end_color='FFFF99',
                                                                   fill_type='solid')
                    sheet.cell(row=9, column=1).fill = PatternFill(start_color='FFFF99', end_color='FFFF99',
                                                                   fill_type='solid')
                    sheet.cell(row=15, column=1).fill = PatternFill(start_color='FFFF99', end_color='FFFF99',
                                                                    fill_type='solid')
                    sheet.cell(row=22, column=1).fill = PatternFill(start_color='FFFF99', end_color='FFFF99',
                                                                    fill_type='solid')

                    sheet1.cell(row=2, column=1).fill = PatternFill(start_color='FFFF99', end_color='FFFF99',
                                                                    fill_type='solid')
                    if (flr == 1):
                        pass
                    else:
                        sheet.cell(row=26, column=1).fill = PatternFill(start_color='FFFF99', end_color='FFFF99',
                                                                        fill_type='solid')

                    # -------------------------------------------------
                    #      Cell Formatting END
                    # -------------------------------------------------

                    # -------------------------------------------------
                    #       Remove Not USED Sheet
                    # -------------------------------------------------

                    wb.remove_sheet(wb.get_sheet_by_name('DETAIL-CUMULUS-BOM-MEDIUM-SITE'))
                    wb.remove_sheet(wb.get_sheet_by_name('CUMULUS Large-Site-Type'))
                    wb.remove_sheet(wb.get_sheet_by_name('CISCO Small-Site-Type'))
                    wb.remove_sheet(wb.get_sheet_by_name('CISCO Medium-Site-Type'))
                    wb.remove_sheet(wb.get_sheet_by_name('CISCO Large-Site-Type'))

                    # -------------------------------------------------
                    #      Removed Not USED Sheet Code END
                    # -------------------------------------------------

                    wb.save(s)

                    # -------------------------------------------------
                    #       CODE FOR MAIL STARTS
                    # -------------------------------------------------

                '''subject = "An email with attachment from Python"
                    body = "This is an email with attachment sent from Python"
                    sender_email = "pythontesting13june@gmail.com"
                    receiver_email = "er.akash.dhand@gmail.com"
                    password = "python@1234"

                    # Create a multipart message and set headers
                    message = MIMEMultipart()
                    message["From"] = sender_email
                    message["To"] = receiver_email
                    message["Subject"] = subject
                    message["Bcc"] = receiver_email  # Recommended for mass emails

                    # Add body to email
                    message.attach(MIMEText(body, "plain"))

                    filename = 'result{}.xlsx'.format(now.strftime("%c")).replace(':','_')  # In same directory as script

                    # Open PDF file in binary mode
                    with open(filename, "rb") as attachment:
                        # Add file as application/octet-stream
                        # Email client can usually download this automatically as attachment
                        part = MIMEBase("application", "octet-stream")
                        part.set_payload(attachment.read())

                    # Encode file in ASCII characters to send by email
                    encoders.encode_base64(part)

                    # Add header as key/value pair to attachment part
                    part.add_header(
                        "Content-Disposition",
                        f"attachment; filename= {filename}",
                    )

                    # Add attachment to message and convert message to string
                    message.attach(part)
                    text = message.as_string()

                    # Log in to server using secure context and send email
                    context = ssl.create_default_context()
                    with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=context) as server:
                        server.login(sender_email, password)
                        server.sendmail(sender_email, receiver_email, text)

                    # -------------------------------------------------
                    #       CODE FOR MAIL ENDS
                    # -------------------------------------------------
                '''

                # -------------------------------------------------
                #       XL function ends
                # -------------------------------------------------

                def XL_CUMULUS_Medium_Site_Type():
                    start_row = 0
                    start_row111 = 0
                    start_row1 = 0
                    start_row2 = 0
                    start_row3 = 0
                    start_row4 = 0
                    start_row5 = 0
                    end_row = 1
                    end_row12 = 1
                    end_row22 = 1
                    end_row32 = 1
                    end_row42 = 1
                    # -------------------------------------------------
                    #       CODE FOR creating user access layer and core layer Dictionaries
                    # -------------------------------------------------
                    wb = xlrd.open_workbook('BOM Code.xls')
                    sh = wb.sheet_by_name('DETAIL-CUMULUS-BOM-MEDIUM-SITE')
                    num_rows = sh.nrows
                    num_cells = sh.ncols - 1
                    sh1 = wb.sheet_by_name('SUMMARY-CUMULUS-BOM')
                    num_rows1 = sh1.nrows
                    num_cells1 = sh1.ncols - 1

                    curr_row = 0

                    # -------------------------------------------------
                    #       CODE FOR writing back to xl
                    # -------------------------------------------------
                    print("Medum xl created")
                    rb = xlrd.open_workbook("BOM Code.xls")
                    sh = rb.sheet_by_name('DETAIL-CUMULUS-BOM-MEDIUM-SITE')
                    sh1 = rb.sheet_by_name('SUMMARY-CUMULUS-BOM')
                    wb = copy(rb)
                    w_sheet = wb.get_sheet('DETAIL-CUMULUS-BOM-MEDIUM-SITE')
                    w_sheet1 = wb.get_sheet('SUMMARY-CUMULUS-BOM')

                    # -------------------------------------------------
                    #       CODE TO GET VALUES INTO CUMULUS-SUMMARY-TAB
                    # -------------------------------------------------
                    # SUMMARY-PAGE

                    for i in range(2, num_rows1):
                        partnumber = sh1.cell(i, 2).value
                        if partnumber == '407-BBOU-US':
                            if flr == 1:
                                v = leaf_list1 * 4 + 46 + flr * 4
                                w_sheet1.write(i, 4, v)
                            else:
                                v = leaf_list1 * 4 + 46 + flr * 4 + 6
                                w_sheet1.write(i, 4, v)

                        if partnumber == '407-BBOS-US':
                            v = 80
                            w_sheet1.write(i, 4, v)
                        if partnumber == '210-ADUX':
                            if flr == 1:
                                v = 5  # (Super Spine A and B +1 extra, SVC spine A and B )
                                w_sheet1.write(i, 4, v)
                            else:
                                v1 = 5 + 2  # (Super Spine A and B +1 extra, SVC spine A and B, Mgmt A and B)
                                w_sheet1.write(i, 4, v1)
                        if partnumber == '4610-54P-O-AC-F-US':
                            v = leaf_list1 + 1 + flr
                            w_sheet1.write(i, 4, v)
                        if partnumber == 'A9748229':

                            if flr == 1:
                                v1 = 2 + 2
                                w_sheet1.write(i, 4, v1)
                            else:
                                v2 = 2 + 2 + 2
                                w_sheet1.write(i, 4, v2)
                        if partnumber == 'A8793201':
                            v = leaf_list1 + flr
                            w_sheet1.write(i, 4, v)

                        if partnumber == 'SLC80482201S':
                            if flr == 1:
                                v = flr
                                w_sheet1.write(i, 4, v)
                            else:
                                v = flr + 2
                                w_sheet1.write(i, 4, v)

                    # -------------------------------------------------
                    #       CODE TO GET VALUES OF ROW and insert leaf value inside qty column
                    # -------------------------------------------------
                    # CUMULUS SPINE ENTRY
                    for i in range(2, num_rows):
                        partnumber = sh.cell(i, 0).value
                        if partnumber == "":
                            end_row = i
                            break
                    Spine_end_row = end_row
                    for i in range(2, Spine_end_row + 1):
                        partnumber = sh.cell(i, 2).value
                        if partnumber == '407-BBOU-US':
                            v = leaf_list1 * 2 + 30
                            w_sheet.write(i, 4, v)
                        if partnumber == '407-BBOS-US':
                            v = 20
                            w_sheet.write(i, 4, v)
                        if partnumber == '210-ADUX':
                            v = 2 + 1
                            w_sheet.write(i, 4, v)
                        if partnumber == 'A9748229':
                            v = 2
                            w_sheet.write(i, 4, v)

                    # CUMULUS SVC SPINE LAYER ENTRY
                    for i in range(2, num_rows):
                        for j in range(0, num_cells):
                            partnumber = sh.cell(i, j).value
                            if partnumber == 'SVC SPINE':
                                start_row = i
                                break
                    start_row1 = start_row
                    for i in range(start_row1, num_rows):
                        partnumber = sh.cell(i, 0).value
                        if partnumber == "":
                            end_row12 = i
                            break
                    SVC_SPINE_end_row = end_row12

                    for i in range(start_row1, SVC_SPINE_end_row + 1):
                        partnumber = sh.cell(i, 2).value
                        if partnumber == '407-BBOU-US':
                            v = 14
                            w_sheet.write(i, 4, v)
                        if partnumber == '407-BBOS-US':
                            v = 60
                            w_sheet.write(i, 4, v)
                        if partnumber == '210-ADUX':
                            v = 2
                            w_sheet.write(i, 4, v)

                        if partnumber == 'A9748229':
                            v = 2
                            w_sheet.write(i, 4, v)

                    # USER Access Leave ENTRY
                    for i in range(2, num_rows):
                        for j in range(0, num_cells):
                            partnumber = sh.cell(i, j).value
                            if partnumber == 'User Access Leaves':
                                start_row = i
                                break
                    start_row2 = start_row

                    for i in range(start_row2, num_rows):
                        partnumber = sh.cell(i, 0).value
                        if partnumber == "":
                            end_row22 = i
                            break
                    User_Access_end_row = end_row22

                    for i in range(start_row2, User_Access_end_row + 1):
                        partnumber = sh.cell(i, 2).value
                        if partnumber == '407-BBOU-US':
                            vv = leaf_list1 * 2 + 2
                            w_sheet.write(i, 4, vv)

                        if partnumber == '4610-54P-O-AC-F-US':
                            v = leaf_list1 + 1
                            w_sheet.write(i, 4, v)
                        if partnumber == 'A8793201':
                            v = leaf_list1
                            w_sheet.write(i, 4, v)

                    # Core OOB MGMT LEAVE ENTRY
                    if flr == 1:
                        for i in range(2, num_rows):
                            for j in range(0, num_cells):
                                partnumber = sh.cell(i, j).value
                                if partnumber == 'Management Core Layer':
                                    start_row = i
                                    break
                        start_row1 = start_row

                        for i in range(start_row1, num_rows):
                            partnumber = sh.cell(i, 0).value
                            if partnumber == "":
                                end_row12 = i
                                break
                        OOB_CORE_end_row = end_row12
                        for i in range(start_row1, OOB_CORE_end_row):
                            partnumber = sh.cell(i, 2).value
                            if partnumber == '407-BBOU-US':
                                v = flr * 2 + 2
                                w_sheet.write(i, 4, v)
                            if partnumber == '407-BBOS-US':
                                v = 4
                                w_sheet.write(i, 4, v)
                            if partnumber == '210-ADUX':
                                v = 1
                                w_sheet.write(i, 4, v)

                                b = 'EDGECORE NETWORKS'
                                w_sheet.write(i, 1, b)
                                c = '4610-54P-O-AC-F-US'
                                w_sheet.write(i, 2, c)
                                d = 'AS4610-54P 48 PORT 10/100/1000BASE-T SWITCH WITH 48 POE+ PORTS,1-8 PORT SUPPORT'
                                w_sheet.write(i, 3, d)
                            if partnumber == 'A9748229':
                                v = 1
                                w_sheet.write(i, 4, v)
                                c1 = 'A8793201'
                                w_sheet.write(i, 2, c1)
                                d1 = 'Cumulus Linux Perpetual License, 1G, 1 Year Software Updates and Support Included'
                                w_sheet.write(i, 3, d1)

                        # Terminal Server Entry
                        for i in range(2, num_rows):
                            for j in range(0, num_cells):
                                partnumber = sh.cell(i, j).value
                                if partnumber == 'Terminal Server':
                                    start_row = i
                                    break

                        start_row1 = start_row
                        for i in range(start_row1, num_rows):
                            partnumber = sh.cell(i, 0).value
                            if partnumber == "":
                                end_row12 = i
                                break
                        OOB_CORE_end_row = end_row12
                        for i in range(start_row1, OOB_CORE_end_row):
                            partnumber = sh.cell(i, 2).value
                            if partnumber == 'SLC80482201S':
                                v = flr
                                w_sheet.write(i, 4, v)
                        # MGMT Access LAYER REMOVAL

                        for i in range(2, num_rows):
                            for j in range(0, num_cells):
                                partnumber = sh.cell(i, j).value
                                if partnumber == 'Management Access Leaves':
                                    start_row = i
                                    break
                        start_row1 = start_row

                        for i in range(OOB_CORE_end_row, num_rows):
                            for j in range(num_cells + 1):
                                w_sheet.write(i, j, "")

                    else:
                        for i in range(2, num_rows):
                            for j in range(0, num_cells):
                                partnumber = sh.cell(i, j).value
                                if partnumber == 'Management Core Layer':
                                    start_row = i
                                    break
                        start_row1 = start_row

                        for i in range(start_row1, num_rows):
                            partnumber = sh.cell(i, 0).value
                            if partnumber == "":
                                end_row12 = i
                                break
                        OOB_CORE_end_row = end_row12
                        for i in range(start_row1, OOB_CORE_end_row):
                            partnumber = sh.cell(i, 2).value
                            if partnumber == '407-BBOU-US':
                                v = flr * 2 + 6
                                w_sheet.write(i, 4, v)
                            if partnumber == '407-BBOS-US':
                                v = 4
                                w_sheet.write(i, 4, v)
                            if partnumber == '210-ADUX':
                                v = 2
                                w_sheet.write(i, 4, v)

                            if partnumber == 'A9748229':
                                v = 2
                                w_sheet.write(i, 4, v)
                        # Terminal Server Entry
                        for i in range(2, num_rows):
                            for j in range(0, num_cells):
                                partnumber = sh.cell(i, j).value
                                if partnumber == 'Terminal Server':
                                    start_row = i
                                    break

                        start_row1 = start_row
                        for i in range(start_row1, num_rows):
                            partnumber = sh.cell(i, 0).value
                            if partnumber == "":
                                end_row12 = i
                                break
                        OOB_CORE_end_row = end_row12
                        for i in range(start_row1, OOB_CORE_end_row):
                            partnumber = sh.cell(i, 2).value
                            if partnumber == 'SLC80482201S':
                                v = flr + 2
                                w_sheet.write(i, 4, v)
                        # MGMT Leaves LAYER ENTRY
                        for i in range(2, num_rows):
                            for j in range(0, num_cells):
                                partnumber = sh.cell(i, j).value
                                if partnumber == 'Management Access Leaves':
                                    start_row = i
                                    break
                        start_row1 = start_row

                        for i in range(start_row1, num_rows):

                            partnumber = sh.cell(i, 0).value
                            if partnumber == "":
                                end_row12 = i
                                break
                        OOB_Leaves_end_row = end_row12
                        for i in range(start_row1, num_rows):
                            partnumber = sh.cell(i, 2).value
                            if partnumber == '407-BBOU-US':
                                v = flr * 2
                                w_sheet.write(i, 4, v)

                            if partnumber == '4610-54P-O-AC-F-US' or partnumber == 'A8793201':
                                v = flr
                                w_sheet.write(i, 4, v)
                    # -------------------------------------------------
                    #       CODE FOR saving xl file after updating
                    # -------------------------------------------------
                    now = datetime.now()
                    s = str('result{}.xls'.format(now.strftime("%c")))
                    wb.save(s.replace(':', '_'))
                    pyexcel.save_book_as(file_name=s.replace(':', '_'),
                                         dest_file_name='result{}.xlsx'.format(now.strftime("%c")).replace(':',
                                                                                                           '_'))
                    s = 'result{}.xlsx'.format(now.strftime("%c")).replace(':', '_')
                    # -------------------------------------------------
                    #       Used openpyxl for editing the xlsx file
                    # -------------------------------------------------
                    import openpyxl
                    from openpyxl.styles import Font
                    wb = openpyxl.load_workbook(s)
                    sheet = wb['DETAIL-CUMULUS-BOM-MEDIUM-SITE']
                    sheet1 = wb['SUMMARY-CUMULUS-BOM']
                    # -------------------------------------------------
                    #       CODE FOR col width(adjust the col width according to text length in it)
                    # -------------------------------------------------
                    dims = {}
                    for row in sheet.rows:
                        for cell in row:
                            if cell.value:
                                dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))
                    for col, value in dims.items():
                        sheet.column_dimensions[col].width = value

                    sheet.column_dimensions['E'].width = 10
                    sheet1.column_dimensions['A'].width = 15
                    sheet1.column_dimensions['B'].width = 25
                    sheet1.column_dimensions['C'].width = 25
                    sheet1.column_dimensions['D'].width = 110
                    sheet1.column_dimensions['E'].width = 10

                    # -------------------------------------------------
                    #      Cell Formatting
                    # -------------------------------------------------

                    top_left11_cell = sheet1['A2']

                    top_left_cell = sheet['A2']
                    top_left1_cell = sheet['A9']
                    top_left2_cell = sheet['A16']
                    top_left3_cell = sheet['A22']
                    top_left4_cell = sheet['A28']
                    top_left5_cell = sheet['A32']
                    sheet.merge_cells('A2:E2')
                    sheet.merge_cells('A9:E9')
                    sheet.merge_cells('A16:E16')
                    sheet.merge_cells('A22:E22')
                    sheet.merge_cells('A28:E28')
                    sheet.merge_cells('A32:E32')

                    sheet1.merge_cells('A2:E2')

                    top_left_cell.alignment = Alignment(horizontal="center", vertical="center")
                    top_left1_cell.alignment = Alignment(horizontal="center", vertical="center")
                    top_left2_cell.alignment = Alignment(horizontal="center", vertical="center")
                    top_left3_cell.alignment = Alignment(horizontal="center", vertical="center")
                    top_left4_cell.alignment = Alignment(horizontal="center", vertical="center")
                    top_left5_cell.alignment = Alignment(horizontal="center", vertical="center")

                    top_left11_cell.alignment = Alignment(horizontal="center", vertical="center")

                    sheet.cell(row=2, column=1).fill = PatternFill(start_color='FFFF99', end_color='FFFF99',
                                                                   fill_type='solid')
                    sheet.cell(row=9, column=1).fill = PatternFill(start_color='FFFF99', end_color='FFFF99',
                                                                   fill_type='solid')
                    sheet.cell(row=16, column=1).fill = PatternFill(start_color='FFFF99', end_color='FFFF99',
                                                                    fill_type='solid')

                    sheet.cell(row=22, column=1).fill = PatternFill(start_color='FFFF99', end_color='FFFF99',
                                                                    fill_type='solid')

                    sheet.cell(row=28, column=1).fill = PatternFill(start_color='FFFF99', end_color='FFFF99',
                                                                    fill_type='solid')

                    sheet1.cell(row=2, column=1).fill = PatternFill(start_color='FFFF99', end_color='FFFF99',
                                                                    fill_type='solid')
                    if flr == 1:
                        pass
                    else:
                        sheet.cell(row=32, column=1).fill = PatternFill(start_color='FFFF99', end_color='FFFF99',
                                                                        fill_type='solid')

                    # -------------------------------------------------
                    #      Cell Formatting END
                    # -------------------------------------------------

                    # -------------------------------------------------
                    #       Remove Not USED Sheet
                    # -------------------------------------------------
                    wb.remove_sheet(wb.get_sheet_by_name('DETAIL-CUMULUS-BOM-SMALL-SITE'))
                    wb.remove_sheet(wb.get_sheet_by_name('CUMULUS Large-Site-Type'))
                    wb.remove_sheet(wb.get_sheet_by_name('CISCO Small-Site-Type'))
                    wb.remove_sheet(wb.get_sheet_by_name('CISCO Medium-Site-Type'))
                    wb.remove_sheet(wb.get_sheet_by_name('CISCO Large-Site-Type'))

                    # -------------------------------------------------
                    #      Removed Not USED Sheet Code END
                    # -------------------------------------------------

                    wb.save(s)

                    # -------------------------------------------------
                    #       CODE FOR MAIL STARTS
                    # -------------------------------------------------
                    '''subject = "An email with attachment from Python"
                    body = "This is an email with attachment sent from Python"
                    sender_email = "pythontesting13june@gmail.com"
                    receiver_email = "er.akash.dhand@gmail.com"
                    password = "python@1234"

                    # Create a multipart message and set headers
                    message = MIMEMultipart()
                    message["From"] = sender_email
                    message["To"] = receiver_email
                    message["Subject"] = subject
                    message["Bcc"] = receiver_email  # Recommended for mass emails

                    # Add body to email
                    message.attach(MIMEText(body, "plain"))

                    filename = 'result{}.xlsx'.format(now.strftime("%c")).replace(':','_')  # In same directory as script

                    # Open PDF file in binary mode
                    with open(filename, "rb") as attachment:
                        # Add file as application/octet-stream
                        # Email client can usually download this automatically as attachment
                        part = MIMEBase("application", "octet-stream")
                        part.set_payload(attachment.read())

                    # Encode file in ASCII characters to send by email
                    encoders.encode_base64(part)

                    # Add header as key/value pair to attachment part
                    part.add_header(
                        "Content-Disposition",
                        f"attachment; filename= {filename}",
                    )

                    # Add attachment to message and convert message to string
                    message.attach(part)
                    text = message.as_string()

                    # Log in to server using secure context and send email
                    context = ssl.create_default_context()
                    with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=context) as server:
                        server.login(sender_email, password)
                        server.sendmail(sender_email, receiver_email, text)

                    # -------------------------------------------------
                    #       CODE FOR MAIL ENDS
                    # -------------------------------------------------
                '''

                # -------------------------------------------------
                #       XL function ends
                # -------------------------------------------------
                # ---------------------------------------------------------
                #           Code for large excel site(Starts)
                # ---------------------------------------------------------
                def XL_CUMULUS_Large_Site_Type():
                    start_row = 0
                    end_row = 1
                    end_row12 = 0
                    start_row_dist = 0
                    start_row_access = 0
                    start_row_service = 0
                    end_row_dist = 0

                    # -------------------------------------------------
                    #       CODE FOR creating user access layer and core layer Dictionaries
                    # -------------------------------------------------

                    wb = xlrd.open_workbook('BOM Code.xls')
                    sh = wb.sheet_by_name('CUMULUS Large-Site-Type')
                    num_rows = sh.nrows
                    num_cells = sh.ncols - 1

                    sh1 = wb.sheet_by_name('SUMMARY-CUMULUS-BOM')
                    num_rows1 = sh1.nrows
                    num_cells1 = sh1.ncols - 1

                    curr_row = 0
                    # -------------------------------------------------
                    #       CODE FOR writing back to xl
                    # -------------------------------------------------
                    print("xl created")
                    rb = xlrd.open_workbook("BOM Code.xls")
                    sh = rb.sheet_by_name('CUMULUS Large-Site-Type')
                    sh1 = rb.sheet_by_name('SUMMARY-CUMULUS-BOM')
                    wb = copy(rb)
                    w_sheet = wb.get_sheet('CUMULUS Large-Site-Type')
                    w_sheet1 = wb.get_sheet('SUMMARY-CUMULUS-BOM')

                    # -------------------------------------------------
                    #       CODE TO GET VALUES INTO CUMULUS-SUMMARY-TAB
                    # -------------------------------------------------
                    # SUMMARY-PAGE

                    for i in range(2, num_rows1):
                        partnumber = sh1.cell(i, 2).value
                        if partnumber == '407-BBOU-US':
                            if flr == 1:
                                if leaf_list1 >= 1 and leaf_list1 <= 40:
                                    v = 2 * 2 + leaf_list1 * 4 + 46 + 14 + flr * 2  # (2 spine)*2+ Leaf*4+ 30(for Super spine)+16(for Spine) Extra)+ 14 for SVC spine
                                    w_sheet1.write(i, 4, v)
                                if leaf_list1 >= 41 and leaf_list1 <= 80:
                                    v = 4 * 2 + leaf_list1 * 4 + 62 + 14 + flr * 2  # (4 spine)*2+ Leaf*4+ 30(for Super spine)+32(for Spine) Extra)+14 for SVC spine
                                    w_sheet1.write(i, 4, v)
                                if leaf_list1 >= 81 and leaf_list1 <= 120:
                                    v = 6 * 2 + leaf_list1 * 4 + 78 + 14 + flr * 2  # (6 spine)*2+ Leaf*4+ 30(for Super spine)+48(for Spine) Extra)+14 for SVC spine
                                    w_sheet1.write(i, 4, v)
                                if leaf_list1 >= 121 and leaf_list1 <= 160:
                                    v = 8 * 2 + leaf_list1 * 4 + 94 + 14 + flr * 2  # (8 spine)*2+ Leaf*4+ 30(for Super spine)+64(for Spine) Extra)+14 for SVC spine
                                    w_sheet1.write(i, 4, v)
                                if leaf_list1 >= 161 and leaf_list1 <= 200:
                                    v = 10 * 2 + leaf_list1 * 4 + 110 + 14 + flr * 2  # (10 spine)*2+ Leaf*4+ 30(for Super spine)+80(for Spine) Extra)+14 for SVC spine
                                    w_sheet1.write(i, 4, v)
                                if leaf_list1 >= 201 and leaf_list1 <= 240:
                                    v = 12 * 2 + leaf_list1 * 4 + 126 + 14 + flr * 2  # (12 spine)*2+ Leaf*4+ 30(for Super spine)+96(for Spine) Extra)+14 for SVC spine
                                    w_sheet1.write(i, 4, v)
                                if leaf_list1 >= 241 and leaf_list1 <= 280:
                                    v = 14 * 2 + leaf_list1 * 4 + 142 + 14 + flr * 2  # (14 spine)*2+ Leaf*4+ 30(for Super spine)+112(for Spine) Extra)+14 for SVC spine
                                    w_sheet1.write(i, 4, v)
                                if leaf_list1 >= 281 and leaf_list1 <= 320:
                                    v = 16 * 2 + leaf_list1 * 4 + 158 + 14 + flr * 2  # (16 spine)*2+ Leaf*4+ 30(for Super spine)+128(for Spine) Extra)+14 for SVC spine
                                    w_sheet1.write(i, 4, v)
                                if leaf_list1 >= 321 and leaf_list1 <= 360:
                                    v = 18 * 2 + leaf_list1 * 4 + 174 + 14 + flr * 2  # (18 spine)*2+ Leaf*4+ 30(for Super spine)+144(for Spine) Extra)+14 for SVC spine
                                    w_sheet1.write(i, 4, v)
                                if leaf_list1 >= 361 and leaf_list1 <= 400:
                                    v = 20 * 2 + leaf_list1 * 4 + 190 + 14 + flr * 2  # (20 spine)*2+ Leaf*4+ 30(for Super spine)+160(for Spine) Extra)+14 for SVC spine
                                    w_sheet1.write(i, 4, v)
                                if leaf_list1 >= 401 and leaf_list1 <= 440:
                                    v = 22 * 2 + leaf_list1 * 4 + 206 + 14 + flr * 2  # (20 spine)*2+ Leaf*4+ 30(for Super spine)+176(for Spine) Extra)+14 for SVC spine
                                    w_sheet1.write(i, 4, v)
                            else:
                                if leaf_list1 >= 1 and leaf_list1 <= 40:
                                    v = 2 * 2 + leaf_list1 * 4 + 46 + 14 + flr * 4 + 6  # (2 spine)*2+ Leaf*4+ 30(for Super spine)+16(for Spine) Extra)+ 14 for SVC spine
                                    w_sheet1.write(i, 4, v)
                                if leaf_list1 >= 41 and leaf_list1 <= 80:
                                    v = 4 * 2 + leaf_list1 * 4 + 62 + 14 + flr * 4 + 6  # (4 spine)*2+ Leaf*4+ 30(for Super spine)+32(for Spine) Extra)+14 for SVC spine
                                    w_sheet1.write(i, 4, v)
                                if leaf_list1 >= 81 and leaf_list1 <= 120:
                                    v = 6 * 2 + leaf_list1 * 4 + 78 + 14 + flr * 4 + 6  # (6 spine)*2+ Leaf*4+ 30(for Super spine)+48(for Spine) Extra)+14 for SVC spine
                                    w_sheet1.write(i, 4, v)
                                if leaf_list1 >= 121 and leaf_list1 <= 160:
                                    v = 8 * 2 + leaf_list1 * 4 + 94 + 14 + flr * 4 + 6  # (8 spine)*2+ Leaf*4+ 30(for Super spine)+64(for Spine) Extra)+14 for SVC spine
                                    w_sheet1.write(i, 4, v)
                                if leaf_list1 >= 161 and leaf_list1 <= 200:
                                    v = 10 * 2 + leaf_list1 * 4 + 110 + 14 + flr * 4 + 6  # (10 spine)*2+ Leaf*4+ 30(for Super spine)+80(for Spine) Extra)+14 for SVC spine
                                    w_sheet1.write(i, 4, v)
                                if leaf_list1 >= 201 and leaf_list1 <= 240:
                                    v = 12 * 2 + leaf_list1 * 4 + 126 + 14 + flr * 4 + 6  # (12 spine)*2+ Leaf*4+ 30(for Super spine)+96(for Spine) Extra)+14 for SVC spine
                                    w_sheet1.write(i, 4, v)
                                if leaf_list1 >= 241 and leaf_list1 <= 280:
                                    v = 14 * 2 + leaf_list1 * 4 + 142 + 14 + flr * 4 + 6  # (14 spine)*2+ Leaf*4+ 30(for Super spine)+112(for Spine) Extra)+14 for SVC spine
                                    w_sheet1.write(i, 4, v)
                                if leaf_list1 >= 281 and leaf_list1 <= 320:
                                    v = 16 * 2 + leaf_list1 * 4 + 158 + 14 + flr * 4 + 6  # (16 spine)*2+ Leaf*4+ 30(for Super spine)+128(for Spine) Extra)+14 for SVC spine
                                    w_sheet1.write(i, 4, v)
                                if leaf_list1 >= 321 and leaf_list1 <= 360:
                                    v = 18 * 2 + leaf_list1 * 4 + 174 + 14 + flr * 4 + 6  # (18 spine)*2+ Leaf*4+ 30(for Super spine)+144(for Spine) Extra)+14 for SVC spine
                                    w_sheet1.write(i, 4, v)
                                if leaf_list1 >= 361 and leaf_list1 <= 400:
                                    v = 20 * 2 + leaf_list1 * 4 + 190 + 14 + flr * 4 + 6  # (20 spine)*2+ Leaf*4+ 30(for Super spine)+160(for Spine) Extra)+14 for SVC spine
                                    w_sheet1.write(i, 4, v)
                                if leaf_list1 >= 401 and leaf_list1 <= 440:
                                    v = 22 * 2 + leaf_list1 * 4 + 206 + 14 + flr * 4 + 6  # (20 spine)*2+ Leaf*4+ 30(for Super spine)+176(for Spine) Extra)+14 for SVC spine
                                    w_sheet1.write(i, 4, v)

                        if partnumber == '407-BBOS-US':
                            if leaf_list1 >= 1 and leaf_list1 <= 40:
                                v = 4 + 10 + 80 + 4  # 10 For Super Spine +4 for Spine + 80 for SVC spine + 4 for MGT
                                w_sheet1.write(i, 4, v)
                            if leaf_list1 >= 41 and leaf_list1 <= 80:
                                v = 8 + 10 + 80 + 4  # 10 For Super Spine +8 for Spine + 80 for SVC spine + 4 for MGT
                                w_sheet1.write(i, 4, v)
                            if leaf_list1 >= 81 and sum(leaf_list) <= 120:
                                v = 12 + 10 + 80 + 4  # 10 For Super Spine +12 for Spine + 80 for SVC spine + 4 for MGT
                                w_sheet1.write(i, 4, v)
                            if leaf_list1 >= 121 and leaf_list1 <= 160:
                                v = 16 + 10 + 80 + 4  # 10 For Super Spine +16 for Spine + 80 for SVC spine + 4 for MGT
                                w_sheet1.write(i, 4, v)
                            if leaf_list1 >= 161 and leaf_list1 <= 200:
                                v = 20 + 10 + 80 + 4  # 10 For Super Spine +20 for Spine + 80 for SVC spine + 4 for MGT
                                w_sheet1.write(i, 4, v)
                            if leaf_list1 >= 201 and leaf_list1 <= 240:
                                v = 24 + 10 + 80 + 4  # 10 For Super Spine +24 for Spine + 80 for SVC spine + 4 for MGT
                                w_sheet1.write(i, 4, v)
                            if leaf_list1 >= 241 and leaf_list1 <= 280:
                                v = 28 + 10 + 80 + 4  # 10 For Super Spine +28 for Spine + 80 for SVC spine + 4 for MGT
                                w_sheet1.write(i, 4, v)
                            if leaf_list1 >= 281 and leaf_list1 <= 320:
                                v = 32 + 10 + 80 + 4  # 10 For Super Spine +32 for Spine + 80 for SVC spine + 4 for MGT
                                w_sheet1.write(i, 4, v)
                            if leaf_list1 >= 321 and leaf_list1 <= 360:
                                v = 36 + 10 + 80 + 4  # 10 For Super Spine +36 for Spine + 80 for SVC spine + 4 for MGT
                                w_sheet1.write(i, 4, v)
                            if leaf_list1 >= 361 and leaf_list1 <= 400:
                                v = 40 + 10 + 80 + 4  # 10 For Super Spine +40 for Spine + 80 for SVC spine + 4 for MGT
                                w_sheet1.write(i, 4, v)
                            if leaf_list1 >= 401 and leaf_list1 <= 440:
                                v = 44 + 10 + 80 + 4  # 10 For Super Spine +44 for Spine + 80 for SVC spine + 4 for MGT
                                w_sheet1.write(i, 4, v)

                        if partnumber == '210-ADUX':
                            if flr == 1:
                                if leaf_list1 >= 1 and leaf_list1 <= 40:
                                    v = 2 + 2 + 1 + 2  # (2 For Super Spine , Spine A and Spine B +1, 2 SVC spine)
                                    w_sheet1.write(i, 4, v)
                                if leaf_list1 >= 41 and leaf_list1 <= 80:
                                    v = 2 + 4 + 1 + 2  # (2 For Super Spine ,Spine A,b,c,d +1, 2 SVC spine)
                                    w_sheet1.write(i, 4, v)
                                if leaf_list1 >= 81 and leaf_list1 <= 120:
                                    v = 2 + 6 + 1 + 2  # (2 For Super Spine ,Spine A,b,c,d,e,f +1, 2 SVC spine)
                                    w_sheet1.write(i, 4, v)
                                if leaf_list1 >= 121 and leaf_list1 <= 160:
                                    v = 2 + 8 + 1 + 2  # (2 For Super Spine ,Spine A,b,c,d,e,f,g,h +1, 2 SVC spine)
                                    w_sheet1.write(i, 4, v)
                                if leaf_list1 >= 161 and leaf_list1 <= 200:
                                    v = 2 + 10 + 1 + 2  # (2 For Super Spine ,Spine A,b,c,d,e,f,g,h,i,j +1, 2 SVC spine)
                                    w_sheet1.write(i, 4, v)
                                if leaf_list1 >= 201 and leaf_list1 <= 240:
                                    v = 2 + 12 + 1 + 2  # (2 For Super Spine ,Spine A,b,c,d,e,f,g,h,i,j,k,l +1, 2 SVC spine)
                                    w_sheet1.write(i, 4, v)
                                if leaf_list1 >= 241 and leaf_list1 <= 280:
                                    v = 2 + 14 + 1 + 2  # (2 For Super Spine ,Spine A,b,c,d,e,f,g,h,i,j,k,l,m,n +1, 2 SVC spine)
                                    w_sheet1.write(i, 4, v)
                                if leaf_list1 >= 281 and leaf_list1 <= 320:
                                    v = 2 + 16 + 1 + 2  # (2 For Super Spine ,Spine A,b,c,d,e,f,g,h,i,j,k,l,m,n,o,p +1, 2 SVC spine)
                                    w_sheet1.write(i, 4, v)
                                if leaf_list1 >= 321 and leaf_list1 <= 360:
                                    v = 2 + 18 + 1 + 2  # (2 For Super Spine ,Spine A,b,c,d,e,f,g,h,i,j,k,l,m,n,o,p,q,r +1, 2 SVC spine)
                                    w_sheet1.write(i, 4, v)
                                if leaf_list1 >= 361 and leaf_list1 <= 400:
                                    v = 2 + 20 + 1 + 2  # (2 For Super Spine ,Spine A,b,c,d,e,f,g,h,i,j,k,l,m,n,o,p,q,r,s,t +1, 2 SVC spine)
                                    w_sheet1.write(i, 4, v)
                                if leaf_list1 >= 401 and leaf_list1 <= 440:
                                    v = 2 + 22 + 1 + 2  # (2 For Super Spine ,Spine A,b,c,d,e,f,g,h,i,j,k,l,m,n,o,p,q,r,s,t,u,v +1, 2 SVC spine)
                                    w_sheet1.write(i, 4, v)

                            else:
                                if leaf_list1 >= 1 and leaf_list1 <= 40:
                                    v = 2 + 2 + 1 + 2 + 2  # (2 For Super Spine , Spine A and Spine B +1, 2 SVC spine+2 MGT Spine)
                                    w_sheet1.write(i, 4, v)
                                if leaf_list1 >= 41 and leaf_list1 <= 80:
                                    v = 2 + 4 + 1 + 2 + 2  # (2 For Super Spine ,Spine A,b,c,d +1, 2 SVC spine2 MGT Spine)
                                    w_sheet1.write(i, 4, v)
                                if leaf_list1 >= 81 and leaf_list1 <= 120:
                                    v = 2 + 6 + 1 + 2 + 2  # (2 For Super Spine ,Spine A,b,c,d,e,f +1, 2 SVC spine2 MGT Spine)
                                    w_sheet1.write(i, 4, v)
                                if leaf_list1 >= 121 and leaf_list1 <= 160:
                                    v = 2 + 8 + 1 + 2 + 2  # (2 For Super Spine ,Spine A,b,c,d,e,f,g,h +1, 2 SVC spine2 MGT Spine)
                                    w_sheet1.write(i, 4, v)
                                if leaf_list1 >= 161 and leaf_list1 <= 200:
                                    v = 2 + 10 + 1 + 2 + 2  # (2 For Super Spine ,Spine A,b,c,d,e,f,g,h,i,j +1, 2 SVC spine2 MGT Spine)
                                    w_sheet1.write(i, 4, v)
                                if leaf_list1 >= 201 and leaf_list1 <= 240:
                                    v = 2 + 12 + 1 + 2 + 2  # (2 For Super Spine ,Spine A,b,c,d,e,f,g,h,i,j,k,l +1, 2 SVC spine2 MGT Spine)
                                    w_sheet1.write(i, 4, v)
                                if leaf_list1 >= 241 and leaf_list1 <= 280:
                                    v = 2 + 14 + 1 + 2 + 2  # (2 For Super Spine ,Spine A,b,c,d,e,f,g,h,i,j,k,l,m,n +1, 2 SVC spine2 MGT Spine)
                                    w_sheet1.write(i, 4, v)
                                if leaf_list1 >= 281 and leaf_list1 <= 320:
                                    v = 2 + 16 + 1 + 2 + 2  # (2 For Super Spine ,Spine A,b,c,d,e,f,g,h,i,j,k,l,m,n,o,p +1, 2 SVC spine2 MGT Spine)
                                    w_sheet1.write(i, 4, v)
                                if leaf_list1 >= 321 and leaf_list1 <= 360:
                                    v = 2 + 18 + 1 + 2 + 2  # (2 For Super Spine ,Spine A,b,c,d,e,f,g,h,i,j,k,l,m,n,o,p,q,r +1, 2 SVC spine2 MGT Spine)
                                    w_sheet1.write(i, 4, v)
                                if leaf_list1 >= 361 and leaf_list1 <= 400:
                                    v = 2 + 20 + 1 + 2 + 2  # (2 For Super Spine ,Spine A,b,c,d,e,f,g,h,i,j,k,l,m,n,o,p,q,r,s,t +1, 2 SVC spine2 MGT Spine)
                                    w_sheet1.write(i, 4, v)
                                if leaf_list1 >= 401 and leaf_list1 <= 440:
                                    v = 2 + 22 + 1 + 2 + 2  # (2 For Super Spine ,Spine A,b,c,d,e,f,g,h,i,j,k,l,m,n,o,p,q,r,s,t,u,v +1, 2 SVC spine2 MGT Spine)
                                    w_sheet1.write(i, 4, v)
                        if partnumber == '4610-54P-O-AC-F-US':
                            v = leaf_list1 + 1 + flr
                            w_sheet1.write(i, 4, v)
                        if partnumber == 'A9748229':
                            if flr == 1:
                                if leaf_list1 >= 1 and leaf_list1 <= 40:
                                    v = 2 + 2 + 2  # (2 For Super Spine , Spine A and Spine B +1, 2 SVC spine)
                                    w_sheet1.write(i, 4, v)
                                if leaf_list1 >= 41 and leaf_list1 <= 80:
                                    v = 2 + 4 + 2  # (2 For Super Spine ,Spine A,b,c,d +1, 2 SVC spine)
                                    w_sheet1.write(i, 4, v)
                                if leaf_list1 >= 81 and leaf_list1 <= 120:
                                    v = 2 + 6 + 2  # (2 For Super Spine ,Spine A,b,c,d,e,f +1, 2 SVC spine)
                                    w_sheet1.write(i, 4, v)
                                if leaf_list1 >= 121 and leaf_list1 <= 160:
                                    v = 2 + 8 + 2  # (2 For Super Spine ,Spine A,b,c,d,e,f,g,h +1, 2 SVC spine)
                                    w_sheet1.write(i, 4, v)
                                if leaf_list1 >= 161 and leaf_list1 <= 200:
                                    v = 2 + 10 + 2  # (2 For Super Spine ,Spine A,b,c,d,e,f,g,h,i,j +1, 2 SVC spine)
                                    w_sheet1.write(i, 4, v)
                                if leaf_list1 >= 201 and leaf_list1 <= 240:
                                    v = 2 + 12 + 2  # (2 For Super Spine ,Spine A,b,c,d,e,f,g,h,i,j,k,l +1, 2 SVC spine)
                                    w_sheet1.write(i, 4, v)
                                if leaf_list1 >= 241 and leaf_list1 <= 280:
                                    v = 2 + 14 + 2  # (2 For Super Spine ,Spine A,b,c,d,e,f,g,h,i,j,k,l,m,n +1, 2 SVC spine)
                                    w_sheet1.write(i, 4, v)
                                if leaf_list1 >= 281 and leaf_list1 <= 320:
                                    v = 2 + 16 + 2  # (2 For Super Spine ,Spine A,b,c,d,e,f,g,h,i,j,k,l,m,n,o,p +1, 2 SVC spine)
                                    w_sheet1.write(i, 4, v)
                                if leaf_list1 >= 321 and leaf_list1 <= 360:
                                    v = 2 + 18 + 2  # (2 For Super Spine ,Spine A,b,c,d,e,f,g,h,i,j,k,l,m,n,o,p,q,r +1, 2 SVC spine)
                                    w_sheet1.write(i, 4, v)
                                if leaf_list1 >= 361 and leaf_list1 <= 400:
                                    v = 2 + 20 + 2  # (2 For Super Spine ,Spine A,b,c,d,e,f,g,h,i,j,k,l,m,n,o,p,q,r,s,t +1, 2 SVC spine)
                                    w_sheet1.write(i, 4, v)
                                if leaf_list1 >= 401 and leaf_list1 <= 440:
                                    v = 2 + 22 + 2  # (2 For Super Spine ,Spine A,b,c,d,e,f,g,h,i,j,k,l,m,n,o,p,q,r,s,t,u,v +1, 2 SVC spine)
                                    w_sheet1.write(i, 4, v)

                            else:
                                if leaf_list1 >= 1 and leaf_list1 <= 40:
                                    v = 2 + 2 + 2 + 2  # (2 For Super Spine , Spine A and Spine B, 2 SVC spine+2 MGT Spine)
                                    w_sheet1.write(i, 4, v)
                                if leaf_list1 >= 41 and leaf_list1 <= 80:
                                    v = 2 + 4 + 2 + 2  # (2 For Super Spine ,Spine A,b,c,d, 2 SVC spine2 MGT Spine)
                                    w_sheet1.write(i, 4, v)
                                if leaf_list1 >= 81 and leaf_list1 <= 120:
                                    v = 2 + 6 + 2 + 2  # (2 For Super Spine ,Spine A,b,c,d,e,f, 2 SVC spine2 MGT Spine)
                                    w_sheet1.write(i, 4, v)
                                if leaf_list1 >= 121 and leaf_list1 <= 160:
                                    v = 2 + 8 + 2 + 2  # (2 For Super Spine ,Spine A,b,c,d,e,f,g,h, 2 SVC spine2 MGT Spine)
                                    w_sheet1.write(i, 4, v)
                                if leaf_list1 >= 161 and leaf_list1 <= 200:
                                    v = 2 + 10 + 2 + 2  # (2 For Super Spine ,Spine A,b,c,d,e,f,g,h,i,j, 2 SVC spine2 MGT Spine)
                                    w_sheet1.write(i, 4, v)
                                if leaf_list1 >= 201 and leaf_list1 <= 240:
                                    v = 2 + 12 + 2 + 2  # (2 For Super Spine ,Spine A,b,c,d,e,f,g,h,i,j,k,l, 2 SVC spine2 MGT Spine)
                                    w_sheet1.write(i, 4, v)
                                if leaf_list1 >= 241 and leaf_list1 <= 280:
                                    v = 2 + 14 + 2 + 2  # (2 For Super Spine ,Spine A,b,c,d,e,f,g,h,i,j,k,l,m,n, 2 SVC spine2 MGT Spine)
                                    w_sheet1.write(i, 4, v)
                                if leaf_list1 >= 281 and leaf_list1 <= 320:
                                    v = 2 + 16 + 2 + 2  # (2 For Super Spine ,Spine A,b,c,d,e,f,g,h,i,j,k,l,m,n,o,p, 2 SVC spine2 MGT Spine)
                                    w_sheet1.write(i, 4, v)
                                if leaf_list1 >= 321 and leaf_list1 <= 360:
                                    v = 2 + 18 + 2 + 2  # (2 For Super Spine ,Spine A,b,c,d,e,f,g,h,i,j,k,l,m,n,o,p,q,r, 2 SVC spine2 MGT Spine)
                                    w_sheet1.write(i, 4, v)
                                if leaf_list1 >= 361 and leaf_list1 <= 400:
                                    v = 2 + 20 + 2 + 2  # (2 For Super Spine ,Spine A,b,c,d,e,f,g,h,i,j,k,l,m,n,o,p,q,r,s,t, 2 SVC spine2 MGT Spine)
                                    w_sheet1.write(i, 4, v)
                                if leaf_list1 >= 401 and leaf_list1 <= 440:
                                    v = 2 + 22 + 2 + 2  # (2 For Super Spine ,Spine A,b,c,d,e,f,g,h,i,j,k,l,m,n,o,p,q,r,s,t,u,v, 2 SVC spine2 MGT Spine)
                                    w_sheet1.write(i, 4, v)
                        if partnumber == 'A8793201':
                            v = leaf_list1 + flr
                            w_sheet1.write(i, 4, v)
                        if partnumber == 'SLC80482201S':
                            if flr == 1:
                                v = flr
                                w_sheet1.write(i, 4, v)
                            else:
                                v = flr + 2
                                w_sheet1.write(i, 4, v)
                    # ------------------------------------------------------------------------
                    #       CODE TO GET VALUES OF ROW and insert leaf value inside qty column
                    # -----------------------------------------------------------------------
                    # CORE LAYER ENTRY
                    for i in range(2, num_rows):
                        partnumber = sh.cell(i, 0).value
                        if partnumber == "":
                            end_row = i
                            break
                    Spine_end_row = end_row
                    for i in range(2, Spine_end_row + 1):
                        partnumber = sh.cell(i, 2).value
                        if partnumber == '407-BBOU-US':
                            if leaf_list1 >= 1 and leaf_list1 <= 40:
                                v = 2 * 2 + 30  # (Spine*2+30 Extra)
                                w_sheet.write(i, 4, v)
                            if leaf_list1 >= 41 and leaf_list1 <= 80:
                                v = 4 * 2 + 30  # (Spine*2+30 Extra)
                                w_sheet.write(i, 4, v)
                            if leaf_list1 >= 81 and leaf_list1 <= 120:
                                v = 6 * 2 + 30  # (Spine*2+30 Extra)
                                w_sheet.write(i, 4, v)
                            if leaf_list1 >= 121 and leaf_list1 <= 160:
                                v = 8 * 2 + 30  # (Spine*2+30 Extra)
                                w_sheet.write(i, 4, v)
                            if leaf_list1 >= 161 and leaf_list1 <= 200:
                                v = 10 * 2 + 30  # (Spine*2+30 Extra)
                                w_sheet.write(i, 4, v)
                            if leaf_list1 >= 201 and leaf_list1 <= 240:
                                v = 12 * 2 + 30  # (Spine*2+30 Extra)
                                w_sheet.write(i, 4, v)
                            if leaf_list1 >= 241 and leaf_list1 <= 280:
                                v = 14 * 2 + 30  # (Spine*2+30 Extra)
                                w_sheet.write(i, 4, v)
                            if leaf_list1 >= 281 and leaf_list1 <= 320:
                                v = 16 * 2 + 30  # (Spine*2+30 Extra)
                                w_sheet.write(i, 4, v)
                            if leaf_list1 >= 321 and leaf_list1 <= 360:
                                v = 18 * 2 + 30  # (Spine*2+30 Extra)
                                w_sheet.write(i, 4, v)
                            if leaf_list1 >= 361 and leaf_list1 <= 400:
                                v = 20 * 2 + 30  # (Spine*2+30 Extra)
                                w_sheet.write(i, 4, v)
                            if leaf_list1 >= 401 and leaf_list1 <= 440:
                                v = 22 * 2 + 30  # (Spine*2+30 Extra)
                                w_sheet.write(i, 4, v)
                        if partnumber == '407-BBOS-US':
                            v = 10
                            w_sheet.write(i, 4, v)
                        if partnumber == '210-ADUX':
                            v = 2 + 1  # (2 Spine A and B Plus one spare)
                            w_sheet.write(i, 4, v)
                        if partnumber == 'A9748229':
                            v = 2  # (For 2 Spine A and B)
                            w_sheet.write(i, 4, v)
                    # Spine LAYER ENTRY
                    for i in range(2, num_rows):
                        for j in range(0, num_cells):
                            partnumber = sh.cell(i, j).value
                            if partnumber == 'SPINE':
                                start_row_dist = i
                                break
                    print("start", start_row_dist)
                    for i in range(start_row_dist, num_rows):
                        partnumber = sh.cell(i, 0).value
                        if partnumber == "":
                            end_row_dist = i
                            break
                    print("end", end_row_dist)
                    print("sum", sum(leaf_list))
                    for i in range(start_row_dist, end_row_dist + 1):
                        print("Cumulus Large in the loop", i)
                        partnumber = sh.cell(i, 2).value

                        if partnumber == '407-BBOU-US':
                            if leaf_list1 >= 1 and leaf_list1 <= 40:
                                v = leaf_list1 * 2 + 16  # (Leaf*2+16 Extra)
                                w_sheet.write(i, 4, v)
                            if leaf_list1 >= 41 and leaf_list1 <= 80:
                                v = leaf_list1 * 2 + 32  # (Leaf*2+32 Extra)
                                w_sheet.write(i, 4, v)
                            if leaf_list1 >= 81 and leaf_list1 <= 120:
                                v = leaf_list1 * 2 + 48  # (Leaf*2+48 Extra)
                                w_sheet.write(i, 4, v)
                            if leaf_list1 >= 121 and leaf_list1 <= 160:
                                v = leaf_list1 * 2 + 64  # (Leaf*2+64 Extra)
                                w_sheet.write(i, 4, v)
                            if leaf_list1 >= 161 and leaf_list1 <= 200:
                                v = leaf_list1 * 2 + 80  # (Leaf*2+80 Extra)
                                w_sheet.write(i, 4, v)
                            if leaf_list1 >= 201 and leaf_list1 <= 240:
                                v = leaf_list1 * 2 + 96  # (Leaf*2+96 Extra)
                                w_sheet.write(i, 4, v)
                            if leaf_list1 >= 241 and leaf_list1 <= 280:
                                v = leaf_list1 * 2 + 112  # (Leaf*2+112 Extra)
                                w_sheet.write(i, 4, v)
                            if leaf_list1 >= 281 and leaf_list1 <= 320:
                                v = leaf_list1 * 2 + 128  # (Leaf*2+128 Extra)
                                w_sheet.write(i, 4, v)
                            if leaf_list1 >= 321 and leaf_list1 <= 360:
                                v = leaf_list1 * 2 + 144  # (Leaf*2+144 Extra)
                                w_sheet.write(i, 4, v)
                            if leaf_list1 >= 361 and leaf_list1 <= 400:
                                v = leaf_list1 * 2 + 160  # (Leaf*2+160 Extra)
                                w_sheet.write(i, 4, v)
                            if leaf_list1 >= 401 and leaf_list1 <= 440:
                                v = leaf_list1 * 2 + 176  # (Leaf*2+176 Extra)
                                w_sheet.write(i, 4, v)
                        if partnumber == '407-BBOS-US':
                            if leaf_list1 >= 1 and leaf_list1 <= 40:
                                v = 4
                                w_sheet.write(i, 4, v)
                            if leaf_list1 >= 41 and leaf_list1 <= 80:
                                v = 8
                                w_sheet.write(i, 4, v)
                            if leaf_list1 >= 81 and sum(leaf_list) <= 120:
                                v = 12
                                w_sheet.write(i, 4, v)
                            if leaf_list1 >= 121 and leaf_list1 <= 160:
                                v = 16
                                w_sheet.write(i, 4, v)
                            if leaf_list1 >= 161 and leaf_list1 <= 200:
                                v = 20
                                w_sheet.write(i, 4, v)
                            if leaf_list1 >= 201 and leaf_list1 <= 240:
                                v = 24
                                w_sheet.write(i, 4, v)
                            if leaf_list1 >= 241 and leaf_list1 <= 280:
                                v = 28
                                w_sheet.write(i, 4, v)
                            if leaf_list1 >= 281 and leaf_list1 <= 320:
                                v = 32
                                w_sheet.write(i, 4, v)
                            if leaf_list1 >= 321 and leaf_list1 <= 360:
                                v = 36
                                w_sheet.write(i, 4, v)
                            if leaf_list1 >= 361 and leaf_list1 <= 400:
                                v = 40
                                w_sheet.write(i, 4, v)
                            if leaf_list1 >= 401 and leaf_list1 <= 440:
                                v = 44
                                w_sheet.write(i, 4, v)
                        if partnumber == '210-ADUX':
                            if leaf_list1 >= 1 and leaf_list1 <= 40:
                                v = 2  # (Spine A and Spine B +1)
                                w_sheet.write(i, 4, v)
                            if leaf_list1 >= 41 and leaf_list1 <= 80:
                                v = 4  # (Spine A,b,c,d +1)
                                w_sheet.write(i, 4, v)
                            if leaf_list1 >= 81 and leaf_list1 <= 120:
                                v = 6  # (Spine A,b,c,d,e,f +1)
                                w_sheet.write(i, 4, v)
                            if leaf_list1 >= 121 and leaf_list1 <= 160:
                                v = 8  # (Spine A,b,c,d,e,f,g,h +1)
                                w_sheet.write(i, 4, v)
                            if leaf_list1 >= 161 and leaf_list1 <= 200:
                                v = 10  # (Spine A,b,c,d,e,f,g,h,i,j +1)
                                w_sheet.write(i, 4, v)
                            if leaf_list1 >= 201 and leaf_list1 <= 240:
                                v = 12  # (Spine A,b,c,d,e,f,g,h,i,j,k,l +1)
                                w_sheet.write(i, 4, v)
                            if leaf_list1 >= 241 and leaf_list1 <= 280:
                                v = 14  # (Spine A,b,c,d,e,f,g,h,i,j,k,l,m,n +1)
                                w_sheet.write(i, 4, v)
                            if leaf_list1 >= 281 and leaf_list1 <= 320:
                                v = 16  # (Spine A,b,c,d,e,f,g,h,i,j,k,l,m,n,o,p +1)
                                w_sheet.write(i, 4, v)
                            if leaf_list1 >= 321 and leaf_list1 <= 360:
                                v = 18  # (Spine A,b,c,d,e,f,g,h,i,j,k,l,m,n,o,p,q,r +1)
                                w_sheet.write(i, 4, v)
                            if leaf_list1 >= 361 and leaf_list1 <= 400:
                                v = 20  # (Spine A,b,c,d,e,f,g,h,i,j,k,l,m,n,o,p,q,r,s,t +1)
                                w_sheet.write(i, 4, v)
                            if leaf_list1 >= 401 and leaf_list1 <= 440:
                                v = 22  # (Spine A,b,c,d,e,f,g,h,i,j,k,l,m,n,o,p,q,r,s,t,u,v +1)
                                w_sheet.write(i, 4, v)
                        if partnumber == 'A9748229':
                            if leaf_list1 >= 1 and leaf_list1 <= 40:
                                v = 2  # (Spine A and Spine B)
                                w_sheet.write(i, 4, v)
                            if leaf_list1 >= 41 and sum(leaf_list) <= 80:
                                v = 4  # (Spine A,b,c,d )
                                w_sheet.write(i, 4, v)
                            if leaf_list1 >= 81 and leaf_list1 <= 120:
                                v = 6  # (Spine A,b,c,d,e,f)
                                w_sheet.write(i, 4, v)
                            if leaf_list1 >= 121 and leaf_list1 <= 160:
                                v = 8  # (Spine A,b,c,d,e,f,g,h)
                                w_sheet.write(i, 4, v)
                            if leaf_list1 >= 161 and leaf_list1 <= 200:
                                v = 10  # (Spine A,b,c,d,e,f,g,h,i,j)
                                w_sheet.write(i, 4, v)
                            if leaf_list1 >= 201 and leaf_list1 <= 240:
                                v = 12  # (Spine A,b,c,d,e,f,g,h,i,j,k,l )
                                w_sheet.write(i, 4, v)
                            if leaf_list1 >= 241 and leaf_list1 <= 280:
                                v = 14  # (Spine A,b,c,d,e,f,g,h,i,j,k,l,m,n)
                                w_sheet.write(i, 4, v)
                            if leaf_list1 >= 281 and leaf_list1 <= 320:
                                v = 16  # (Spine A,b,c,d,e,f,g,h,i,j,k,l,m,n,o,p)
                                w_sheet.write(i, 4, v)
                            if leaf_list1 >= 321 and leaf_list1 <= 360:
                                v = 18  # (Spine A,b,c,d,e,f,g,h,i,j,k,l,m,n,o,p,q,r)
                                w_sheet.write(i, 4, v)
                            if leaf_list1 >= 361 and leaf_list1 <= 400:
                                v = 20  # (Spine A,b,c,d,e,f,g,h,i,j,k,l,m,n,o,p,q,r,s,t)
                                w_sheet.write(i, 4, v)
                            if leaf_list1 >= 401 and leaf_list1 <= 440:
                                v = 22  # (Spine A,b,c,d,e,f,g,h,i,j,k,l,m,n,o,p,q,r,s,t,u,v)
                                w_sheet.write(i, 4, v)

                    # SVC SPINE LAYER ENTRY
                    for i in range(2, num_rows):
                        for j in range(0, num_cells):
                            partnumber = sh.cell(i, j).value
                            if partnumber == 'SVC SPINE':
                                start_row = i
                                break
                    start_row1 = start_row
                    for i in range(start_row, num_rows):
                        partnumber = sh.cell(i, 0).value
                        if partnumber == "":
                            end_row = i
                            break
                    SVC_SPINE_end_row = end_row

                    for i in range(start_row1, SVC_SPINE_end_row + 1):
                        partnumber = sh.cell(i, 2).value
                        if partnumber == '407-BBOU-US':
                            v = 14
                            w_sheet.write(i, 4, v)
                        if partnumber == '407-BBOS-US':
                            v = 80
                            w_sheet.write(i, 4, v)
                        if partnumber == '210-ADUX':
                            v = 2
                            w_sheet.write(i, 4, v)
                        if partnumber == 'A9748229':
                            v = 2
                            w_sheet.write(i, 4, v)

                    # USER ACCESS LAYER ENTRY
                    for i in range(2, num_rows):
                        for j in range(0, num_cells):
                            partnumber = sh.cell(i, j).value
                            if partnumber == 'User Access Leaves':
                                start_row = i
                                break
                    start_row2 = start_row

                    for j in range(start_row2, num_rows):
                        partnumber = sh.cell(j, 0).value
                        if partnumber == "":
                            end_row12 = j
                            break
                    USER_ACCESS_Layer_end_row = end_row12

                    for i in range(start_row2, USER_ACCESS_Layer_end_row + 1):
                        partnumber = sh.cell(i, 2).value
                        if partnumber == '407-BBOU-US':
                            vv = leaf_list1 * 2
                            w_sheet.write(i, 4, vv)

                        if partnumber == '4610-54P-O-AC-F-US':
                            v = leaf_list1 + 1  # (Leaf count + 1 Spare)
                            w_sheet.write(i, 4, v)
                        if partnumber == 'A8793201':
                            v = leaf_list1
                            w_sheet.write(i, 4, v)

                    # MGMT CORE  LAYER ENTRY

                    if flr == 1:
                        for i in range(2, num_rows):
                            for j in range(0, num_cells):
                                partnumber = sh.cell(i, j).value
                                if partnumber == 'Management Core Layer':
                                    start_row = i
                                    break
                        start_row1 = start_row

                        for i in range(start_row1, num_rows):
                            partnumber = sh.cell(i, 0).value
                            if partnumber == "":
                                end_row12 = i
                                break
                        OOB_CORE_end_row = end_row12
                        for i in range(start_row1, OOB_CORE_end_row):
                            partnumber = sh.cell(i, 2).value
                            if partnumber == '407-BBOU-US':
                                v = flr * 2
                                w_sheet.write(i, 4, v)
                            if partnumber == '407-BBOS-US':
                                v = 4
                                w_sheet.write(i, 4, v)
                            if partnumber == '210-ADUX':
                                v = 1
                                w_sheet.write(i, 4, v)

                                b = 'EDGECORE NETWORKS'
                                w_sheet.write(i, 1, b)
                                c = '4610-54P-O-AC-F-US'
                                w_sheet.write(i, 2, c)
                                d = 'AS4610-54P 48 PORT 10/100/1000BASE-T SWITCH WITH 48 POE+ PORTS,1-8 PORT SUPPORT'
                                w_sheet.write(i, 3, d)

                            if partnumber == 'A9748229':
                                v = 1
                                w_sheet.write(i, 4, v)
                                c1 = 'A8793201'
                                w_sheet.write(i, 2, c1)
                                d1 = 'Cumulus Linux Perpetual License, 1G, 1 Year Software Updates and Support Included'
                                w_sheet.write(i, 3, d1)
                        # Terminal Server Entry
                        for i in range(2, num_rows):
                            for j in range(0, num_cells):
                                partnumber = sh.cell(i, j).value
                                if partnumber == 'Terminal Server':
                                    start_row = i
                                    break
                        start_row1 = start_row
                        for i in range(start_row1, num_rows):
                            partnumber = sh.cell(i, 0).value
                            if partnumber == "":
                                end_row12 = i
                                break

                        OOB_CORE_end_row = end_row12
                        for i in range(start_row1, OOB_CORE_end_row):
                            partnumber = sh.cell(i, 2).value
                            if partnumber == 'SLC80482201S':
                                v = flr
                                w_sheet.write(i, 4, v)

                        # MGMT Access LAYER REMOVAL

                        for i in range(2, num_rows):
                            for j in range(0, num_cells):
                                partnumber = sh.cell(i, j).value
                                if partnumber == 'Management Access Leaves':
                                    start_row = i
                                    break
                        start_row1 = start_row

                        for i in range(OOB_CORE_end_row, num_rows):
                            for j in range(num_cells + 1):
                                w_sheet.write(i, j, "")

                    else:
                        for i in range(2, num_rows):
                            for j in range(0, num_cells):
                                partnumber = sh.cell(i, j).value
                                if partnumber == 'Management Core Layer':
                                    start_row = i
                                    break
                        start_row1 = start_row

                        for i in range(start_row1, num_rows):
                            partnumber = sh.cell(i, 0).value
                            if partnumber == "":
                                end_row12 = i
                                break
                        OOB_CORE_end_row = end_row12
                        for i in range(start_row1, OOB_CORE_end_row):
                            partnumber = sh.cell(i, 2).value
                            if partnumber == '407-BBOU-US':
                                v = flr * 2 + 6
                                w_sheet.write(i, 4, v)
                            if partnumber == '407-BBOS-US':
                                v = 4
                                w_sheet.write(i, 4, v)
                            if partnumber == '210-ADUX':
                                v = 2
                                w_sheet.write(i, 4, v)

                            if partnumber == 'A9748229':
                                v = 2
                                w_sheet.write(i, 4, v)

                        # Terminal Server Entry
                        for i in range(2, num_rows):
                            for j in range(0, num_cells):
                                partnumber = sh.cell(i, j).value
                                if partnumber == 'Terminal Server':
                                    start_row = i
                                    break
                        start_row1 = start_row
                        for i in range(start_row1, num_rows):
                            partnumber = sh.cell(i, 0).value
                            if partnumber == "":
                                end_row12 = i
                                break
                        OOB_CORE_end_row = end_row12
                        for i in range(start_row1, OOB_CORE_end_row):
                            partnumber = sh.cell(i, 2).value
                            if partnumber == 'SLC80482201S':
                                v = flr + 2
                                w_sheet.write(i, 4, v)
                        # MGMT Leaves LAYER ENTRY
                        for i in range(2, num_rows):
                            for j in range(0, num_cells):
                                partnumber = sh.cell(i, j).value
                                if partnumber == 'Management Access Leaves':
                                    start_row = i
                                    break
                        start_row1 = start_row

                        for i in range(start_row1, num_rows):
                            partnumber = sh.cell(i, 2).value
                            if partnumber == '407-BBOU-US':
                                v = flr * 2
                                w_sheet.write(i, 4, v)
                            if partnumber == '4610-54P-O-AC-F-US' or partnumber == 'A8793201':
                                v = flr
                                w_sheet.write(i, 4, v)

                    # -------------------------------------------------
                    #       CODE FOR saving xl file after updating
                    # -------------------------------------------------
                    now = datetime.now()
                    s = str('result{}.xls'.format(now.strftime("%c")))
                    wb.save(s.replace(':', '_'))
                    pyexcel.save_book_as(file_name=s.replace(':', '_'),
                                         dest_file_name='result{}.xlsx'.format(now.strftime("%c")).replace(':',
                                                                                                           '_'))
                    s = 'result{}.xlsx'.format(now.strftime("%c")).replace(':', '_')
                    # -------------------------------------------------
                    #       Used openpyxl for editing the xlsx file
                    # -------------------------------------------------
                    import openpyxl
                    from openpyxl.styles import Font
                    wb = openpyxl.load_workbook(s)
                    sheet = wb['CUMULUS Large-Site-Type']
                    sheet1 = wb['SUMMARY-CUMULUS-BOM']
                    # -------------------------------------------------
                    #       CODE FOR col width(adjust the col width according to text length in it)
                    # -------------------------------------------------
                    dims = {}
                    for row in sheet.rows:
                        for cell in row:
                            if cell.value:
                                dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))
                    for col, value in dims.items():
                        sheet.column_dimensions[col].width = value

                    sheet.column_dimensions['E'].width = 10

                    sheet1.column_dimensions['A'].width = 15
                    sheet1.column_dimensions['B'].width = 25
                    sheet1.column_dimensions['C'].width = 25
                    sheet1.column_dimensions['D'].width = 110
                    sheet1.column_dimensions['E'].width = 10

                    # -------------------------------------------------
                    #      Cell Formatting
                    # -------------------------------------------------

                    top_left11_cell = sheet1['A2']

                    top_left_cell = sheet['A2']
                    top_left1_cell = sheet['A9']
                    top_left2_cell = sheet['A16']
                    top_left3_cell = sheet['A23']
                    top_left4_cell = sheet['A29']
                    top_left5_cell = sheet['A36']
                    top_left6_cell = sheet['A40']
                    sheet.merge_cells('A2:E2')
                    sheet.merge_cells('A9:E9')
                    sheet.merge_cells('A16:E16')
                    sheet.merge_cells('A23:E23')
                    sheet.merge_cells('A29:E29')
                    sheet.merge_cells('A36:E36')
                    sheet.merge_cells('A40:E40')

                    sheet1.merge_cells('A2:E2')

                    top_left_cell.alignment = Alignment(horizontal="center", vertical="center")
                    top_left1_cell.alignment = Alignment(horizontal="center", vertical="center")
                    top_left2_cell.alignment = Alignment(horizontal="center", vertical="center")
                    top_left3_cell.alignment = Alignment(horizontal="center", vertical="center")
                    top_left4_cell.alignment = Alignment(horizontal="center", vertical="center")
                    top_left5_cell.alignment = Alignment(horizontal="center", vertical="center")
                    top_left6_cell.alignment = Alignment(horizontal="center", vertical="center")

                    top_left11_cell.alignment = Alignment(horizontal="center", vertical="center")

                    sheet.cell(row=2, column=1).fill = PatternFill(start_color='FFFF99', end_color='FFFF99',
                                                                   fill_type='solid')
                    sheet.cell(row=9, column=1).fill = PatternFill(start_color='FFFF99', end_color='FFFF99',
                                                                   fill_type='solid')
                    sheet.cell(row=16, column=1).fill = PatternFill(start_color='FFFF99', end_color='FFFF99',
                                                                    fill_type='solid')
                    sheet.cell(row=23, column=1).fill = PatternFill(start_color='FFFF99', end_color='FFFF99',
                                                                    fill_type='solid')
                    sheet.cell(row=29, column=1).fill = PatternFill(start_color='FFFF99', end_color='FFFF99',
                                                                    fill_type='solid')
                    sheet.cell(row=36, column=1).fill = PatternFill(start_color='FFFF99', end_color='FFFF99',
                                                                    fill_type='solid')

                    sheet1.cell(row=2, column=1).fill = PatternFill(start_color='FFFF99', end_color='FFFF99',
                                                                    fill_type='solid')
                    if flr == 1:
                        pass
                    else:
                        sheet.cell(row=40, column=1).fill = PatternFill(start_color='FFFF99', end_color='FFFF99',
                                                                        fill_type='solid')

                    # -------------------------------------------------
                    #      Cell Formatting END
                    # -------------------------------------------------

                    # -------------------------------------------------
                    #       Remove Not USED Sheet
                    # -------------------------------------------------
                    wb.remove_sheet(wb.get_sheet_by_name('DETAIL-CUMULUS-BOM-SMALL-SITE'))
                    wb.remove_sheet(wb.get_sheet_by_name('DETAIL-CUMULUS-BOM-MEDIUM-SITE'))
                    wb.remove_sheet(wb.get_sheet_by_name('CISCO Small-Site-Type'))
                    wb.remove_sheet(wb.get_sheet_by_name('CISCO Medium-Site-Type'))
                    wb.remove_sheet(wb.get_sheet_by_name('CISCO Large-Site-Type'))

                    # -------------------------------------------------
                    #      Removed Not USED Sheet Code END
                    # -------------------------------------------------

                    wb.save(s)

                    # -------------------------------------------------
                    #       CODE FOR MAIL STARTS
                    # -------------------------------------------------

                '''subject = "An email with attachment from Python"
                    body = "This is an email with attachment sent from Python"
                    sender_email = "pythontesting13june@gmail.com"
                    receiver_email = "er.akash.dhand@gmail.com"
                    password = "python@1234"

                    # Create a multipart message and set headers
                    message = MIMEMultipart()
                    message["From"] = sender_email
                    message["To"] = receiver_email
                    message["Subject"] = subject
                    message["Bcc"] = receiver_email  # Recommended for mass emails

                    # Add body to email
                    message.attach(MIMEText(body, "plain"))

                    filename = 'result{}.xlsx'.format(now.strftime("%c")).replace(':','_')  # In same directory as script

                    # Open PDF file in binary mode
                    with open(filename, "rb") as attachment:
                        # Add file as application/octet-stream
                        # Email client can usually download this automatically as attachment
                        part = MIMEBase("application", "octet-stream")
                        part.set_payload(attachment.read())

                    # Encode file in ASCII characters to send by email
                    encoders.encode_base64(part)

                    # Add header as key/value pair to attachment part
                    part.add_header(
                        "Content-Disposition",
                        f"attachment; filename= {filename}",
                    )

                    # Add attachment to message and convert message to string
                    message.attach(part)
                    text = message.as_string()

                    # Log in to server using secure context and send email
                    context = ssl.create_default_context()
                    with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=context) as server:
                        server.login(sender_email, password)
                        server.sendmail(sender_email, receiver_email, text)

                    # -------------------------------------------------
                    #       CODE FOR MAIL ENDS
                    # -------------------------------------------------
                '''
                # -------------------------------------------------
                #       XL function ends
                # -------------------------------------------------
                floor_user_count_list = []
                floor_name_list1 = []

                leaf_list = []

                for i in range(flr):
                    floor_user_count_list.append(per_floor_usr_count_list[i].get())
                    floor_name_list1.append(floor_name_list[i].get())
                floor_user_count_list = [int(i) for i in floor_user_count_list]

                for j in range(len(floor_user_count_list)):
                    leaf_count = math.ceil(floor_user_count_list[j] / 48 + 1)
                    leaf_list.append(leaf_count)

                SUM = 0
                for i in floor_user_count_list:
                    SUM = SUM + i
                leaf_list1 = math.ceil((int(usertext) + SUM) / 48 + 1)
                print("Check count of leaf list", leaf_list1)

                txt = sc.ScrolledText(floor1_txt_frame, width=80, height=10)
                txt.pack()

                txt.insert(tk.INSERT, "Entered AMEX Site-ID :-  {}\n"
                                      "Entered Total Number of floor :- {}\n"
                                      "Entered Total Number of HeadCount:-{} \n"
                                      "Entered Total Number of EUC Devices :- {}\n"
                                      "Entered Country Name :- {}\n"
                                      "Entered Building Name:- {}\n".format(amex_idvar, flr, usertext, SUM,
                                                                            countryVar,
                                                                            buildind))
                txt.insert(tk.INSERT, "\n")

                def New_NGC_Cumulus_Partial_Funtion():

                    Main_page_logo_frame1 = tk.Frame(root, bd=1, relief=tk.GROOVE, bg="black")
                    Main_page_logo_frame1.place(relx=0.1, relwidth=0.3, relheight=1, anchor='n')

                    New_NGC_page_frame = tk.Frame(root, bd=6, relief=tk.SUNKEN)
                    New_NGC_page_frame.place(relx=0.5, rely=0.15, relwidth=0.8, relheight=0.25, anchor='center')

                    NGC_frame_for_floor_input = tk.Frame(root, bd=6, relief=tk.SUNKEN)
                    NGC_frame_for_floor_input.place(x=0.1, y=0.1, relx=0.5, rely=0.4, anchor='n')

                    country = tk.Label(New_NGC_page_frame, text="Enter Country Name          ", font=font)
                    country.grid(row=0, column=1)
                    country_txt = tk.Entry(New_NGC_page_frame, width=30)
                    country_txt.grid(row=0, column=2)

                    region = tk.Label(New_NGC_page_frame, text="Select Region         ", font=font)
                    region.grid(row=1, column=1)

                    combo = ttk1.Combobox(New_NGC_page_frame, width=27)
                    combo['values'] = ('Select ', 'USA/CAN', 'LAC', 'JAPA', 'EMEA')
                    combo.current(0)
                    combo.grid(row=1, column=2)

                    amex_id = tk.Label(New_NGC_page_frame, text="Enter the AMEX Site-ID", font=font)
                    amex_id.grid(row=2, column=1)
                    amex_id_txt = tk.Entry(New_NGC_page_frame, width=30)
                    amex_id_txt.grid(row=2, column=2)

                    building_name = tk.Label(New_NGC_page_frame, text="          Enter The Building Name",
                                             font=font)
                    building_name.grid(row=0, column=4)
                    building_name = tk.Entry(New_NGC_page_frame, width=30)
                    building_name.grid(row=0, column=5)

                    usr_lable = tk.Label(New_NGC_page_frame,
                                         text="          Enter the total Headcount",
                                         font=font)
                    usr_lable.grid(row=1, column=4)
                    usr_txt = tk.Entry(New_NGC_page_frame, width=30)
                    usr_txt.grid(row=1, column=5)

                    flr_lable = tk.Label(New_NGC_page_frame, text="          Total Number of IDF", font=font)
                    flr_lable.grid(row=2, column=4)
                    flr_txt = tk.Entry(New_NGC_page_frame, width=30)
                    flr_txt.grid(row=2, column=5)

                    New_NGC_page_txt_frame111 = tk.Frame(root, bd=6, relief=tk.SUNKEN)
                    New_NGC_page_txt_frame111.place(x=0.1, y=0.1, relx=0.5, rely=0.5, anchor='n')

                    floor1_txt_frame = tk.Frame(root, bd=6, relief=tk.SUNKEN)
                    floor1_txt_frame.place(relx=0.5, rely=0.2, anchor='n')

                    floor2_txt_frame = tk.Frame(root, bd=6, relief=tk.SUNKEN)
                    floor2_txt_frame.place(relx=0.5, rely=0.8, anchor='n')

                    def NGC_CUMULUS_Function():

                        countryVar = country_txt.get()
                        amex_idvar = amex_id_txt.get()
                        usertext = usr_txt.get()
                        buildind = building_name.get()
                        flr = int(flr_txt.get())
                        per_floor_usr_count_list = []
                        floor_name_list = []
                        floor_name_list2 = []

                        for i in range(1, flr + 1):
                            lable = tk.Label(NGC_frame_for_floor_input, text="Enter Floor Number", font=font)
                            lable.grid(row=10 + i, column=1)
                            flr_num_txt = tk.Entry(NGC_frame_for_floor_input, width=30)
                            flr_num_txt.grid(row=10 + i, column=2)
                            floor_name_list.append(flr_num_txt)
                            l = tk.Label(NGC_frame_for_floor_input, text="Enter the EUC count ".format(i),
                                         font=font)
                            l.grid(row=10 + i, column=5)
                            per_flr_usr_count_txt = tk.Entry(NGC_frame_for_floor_input, width=30)
                            per_flr_usr_count_txt.grid(row=10 + i, column=6)
                            per_floor_usr_count_list.append(per_flr_usr_count_txt)

                            def NGC_CUMULUS_Leaf_Calculation():
                                count = False
                                for i in range(len(floor_name_list)):
                                    floor_name_list2.append(floor_name_list[i].get())
                                    print(floor_name_list2)
                                for i1 in floor_name_list2:
                                    if floor_name_list2.count(i1) >= 2:
                                        count = True
                                        print("Count:", count)
                                        break
                                if count == True:
                                    res = MB.askretrycancel('Message title', 'Floor name same')
                                    if res == True:
                                        print('check')
                                        NGC_CUMULUS_Function()
                                else:
                                    # -------------------------------------------------
                                    #       CODE FOR EXCEL FILE
                                    # -------------------------------------------------

                                    def XL_CUMULUS_Small_Site_Type():
                                        start_row = 0
                                        end_row12 = 1
                                        # -------------------------------------------------
                                        #       CODE FOR creating user access layer and core layer Dictionaries
                                        # -------------------------------------------------
                                        wb = xlrd.open_workbook('BOM Code.xls')
                                        sh = wb.sheet_by_name('DETAIL-CUMULUS-BOM-SMALL-SITE')
                                        num_rows = sh.nrows
                                        sh1 = wb.sheet_by_name('SUMMARY-CUMULUS-BOM')
                                        num_rows1 = sh1.nrows

                                        num_cells = sh.ncols - 1
                                        num_cells1 = sh1.ncols - 1
                                        curr_row = 0
                                        # -------------------------------------------------
                                        #       CODE FOR writing back to xl
                                        # -------------------------------------------------
                                        print("xl created")
                                        rb = xlrd.open_workbook("BOM Code.xls")
                                        sh = rb.sheet_by_name('DETAIL-CUMULUS-BOM-SMALL-SITE')
                                        sh1 = rb.sheet_by_name('SUMMARY-CUMULUS-BOM')
                                        wb = copy(rb)
                                        w_sheet = wb.get_sheet('DETAIL-CUMULUS-BOM-SMALL-SITE')
                                        w_sheet1 = wb.get_sheet('SUMMARY-CUMULUS-BOM')

                                        # -------------------------------------------------
                                        #       CODE TO GET VALUES INTO CUMULUS-SUMMARY-TAB
                                        # -------------------------------------------------
                                        # SUMMARY-PAGE
                                        for i in range(2, num_rows1):
                                            partnumber = sh1.cell(i, 2).value
                                            if partnumber == '407-BBOU-US':
                                                if flr == 1:
                                                    v = leaf_list1 * 4 + 22 + 2
                                                    w_sheet1.write(i, 4, v)
                                                else:
                                                    v = leaf_list1 * 4 + 22 + flr * 4 + 6
                                                    w_sheet1.write(i, 4, v)

                                            if partnumber == '407-BBOS-US':
                                                v = 34
                                                w_sheet1.write(i, 4, v)
                                            if partnumber == '210-ADUX':
                                                if flr == 1:
                                                    v = 2 + 1
                                                    w_sheet1.write(i, 4, v)
                                                else:
                                                    v = 2 + 1 + 2
                                                    w_sheet1.write(i, 4, v)

                                            if partnumber == '4610-54P-O-AC-F-US':
                                                v = leaf_list1 + 1 + flr
                                                w_sheet1.write(i, 4, v)

                                            if partnumber == 'A9748229':
                                                if flr == 1:
                                                    v = 2
                                                    w_sheet1.write(i, 4, v)
                                                else:
                                                    v = 4
                                                    w_sheet1.write(i, 4, v)

                                            if partnumber == 'A8793201':
                                                v = leaf_list1 + flr
                                                w_sheet1.write(i, 4, v)

                                            if partnumber == 'SLC80482201S':
                                                if flr == 1:
                                                    v = flr
                                                    w_sheet1.write(i, 4, v)
                                                else:
                                                    v = flr + 2
                                                    w_sheet1.write(i, 4, v)

                                        # -------------------------------------------------
                                        #       CODE TO GET VALUES OF ROW and insert leaf value inside qty column
                                        # -------------------------------------------------
                                        # CORE LAYER ENTRY
                                        for i in range(2, num_rows):
                                            partnumber = sh.cell(i, 0).value
                                            if partnumber == "":
                                                end_row = i
                                                break
                                        core_layer_end_row = end_row

                                        print("end", end_row)

                                        for i in range(2, core_layer_end_row):
                                            partnumber = sh.cell(i, 2).value
                                            if partnumber == '407-BBOU-US':
                                                v = leaf_list1 * 2 + 22
                                                w_sheet.write(i, 4, v)
                                            if partnumber == '407-BBOS-US':
                                                v = 30
                                                w_sheet.write(i, 4, v)
                                            if partnumber == '210-ADUX':
                                                v = 2 + 1
                                                w_sheet.write(i, 4, v)
                                            if partnumber == 'A9748229':
                                                v = 2
                                                w_sheet.write(i, 4, v)

                                        # USER ACCESS LAYER ENTRY
                                        for i in range(2, num_rows):
                                            for j in range(0, num_cells):
                                                partnumber = sh.cell(i, j).value
                                                if partnumber == 'User Access Leaves':
                                                    start_row = i
                                                    break
                                        start_row1 = start_row

                                        for i in range(start_row1, num_rows):
                                            partnumber = sh.cell(i, 0).value
                                            if partnumber == "":
                                                end_row12 = i
                                                break
                                        USR_Access_end_row = end_row12

                                        for i in range(start_row1, USR_Access_end_row):
                                            partnumber = sh.cell(i, 2).value

                                            if partnumber == '407-BBOU-US':
                                                v = leaf_list1 * 2
                                                w_sheet.write(i, 4, v)

                                            if partnumber == '4610-54P-O-AC-F-US':
                                                v = leaf_list1 + 1
                                                w_sheet.write(i, 4, v)
                                            if partnumber == 'A8793201':
                                                v = leaf_list1
                                                w_sheet.write(i, 4, v)

                                        # MGMT CORE  LAYER ENTRY

                                        if flr == 1:
                                            for i in range(2, num_rows):
                                                for j in range(0, num_cells):
                                                    partnumber = sh.cell(i, j).value
                                                    if partnumber == 'Management Core Layer':
                                                        start_row = i
                                                        break
                                            start_row1 = start_row

                                            for i in range(start_row1, num_rows):
                                                partnumber = sh.cell(i, 0).value
                                                if partnumber == "":
                                                    end_row12 = i
                                                    break
                                            OOB_CORE_end_row = end_row12
                                            for i in range(start_row1, OOB_CORE_end_row):
                                                partnumber = sh.cell(i, 2).value
                                                if partnumber == '407-BBOU-US':
                                                    if flr == 1:
                                                        v = flr * 2
                                                        w_sheet.write(i, 4, v)
                                                    else:
                                                        v = flr * 2 + 6
                                                        w_sheet.write(i, 4, v)

                                                if partnumber == '407-BBOS-US':
                                                    v = 4
                                                    w_sheet.write(i, 4, v)
                                                if partnumber == '210-ADUX':
                                                    v = 1
                                                    w_sheet.write(i, 4, v)

                                                    b = 'EDGECORE NETWORKS'
                                                    w_sheet.write(i, 1, b)
                                                    c = '4610-54P-O-AC-F-US'
                                                    w_sheet.write(i, 2, c)
                                                    d = 'AS4610-54P 48 PORT 10/100/1000BASE-T SWITCH WITH 48 POE+ PORTS,1-8 PORT SUPPORT'
                                                    w_sheet.write(i, 3, d)

                                                if partnumber == 'A9748229':
                                                    v = 1
                                                    w_sheet.write(i, 4, v)
                                                    c1 = 'A8793201'
                                                    w_sheet.write(i, 2, c1)
                                                    d1 = 'Cumulus Linux Perpetual License, 1G, 1 Year Software Updates and Support Included'
                                                    w_sheet.write(i, 3, d1)

                                            # Terminal Server Entry
                                            for i in range(2, num_rows):
                                                for j in range(0, num_cells):
                                                    partnumber = sh.cell(i, j).value
                                                    if partnumber == 'Terminal Server':
                                                        start_row = i
                                                        break
                                            start_row1 = start_row

                                            for i in range(start_row1, num_rows):
                                                partnumber = sh.cell(i, 0).value
                                                if partnumber == "":
                                                    end_row12 = i
                                                    break
                                            OOB_CORE_end_row = end_row12
                                            for i in range(start_row1, OOB_CORE_end_row):
                                                partnumber = sh.cell(i, 2).value
                                                if partnumber == 'SLC80482201S':
                                                    v = flr
                                                    w_sheet.write(i, 4, v)

                                            # MGMT Access LAYER REMOVAL

                                            for i in range(2, num_rows):
                                                for j in range(0, num_cells):
                                                    partnumber = sh.cell(i, j).value
                                                    if partnumber == 'Management Access Leaves':
                                                        start_row = i
                                                        break
                                            start_row1 = start_row

                                            for i in range(OOB_CORE_end_row, num_rows):
                                                for j in range(num_cells + 1):
                                                    w_sheet.write(i, j, "")

                                        else:
                                            for i in range(2, num_rows):
                                                for j in range(0, num_cells):
                                                    partnumber = sh.cell(i, j).value
                                                    if partnumber == 'Management Core Layer':
                                                        start_row = i
                                                        break
                                            start_row1 = start_row

                                            for i in range(start_row1, num_rows):
                                                partnumber = sh.cell(i, 0).value
                                                if partnumber == "":
                                                    end_row12 = i
                                                    break
                                            OOB_CORE_end_row = end_row12
                                            for i in range(start_row1, OOB_CORE_end_row):
                                                partnumber = sh.cell(i, 2).value
                                                if partnumber == '407-BBOU-US':
                                                    v = flr * 2 + 6
                                                    w_sheet.write(i, 4, v)
                                                if partnumber == '407-BBOS-US':
                                                    v = 4
                                                    w_sheet.write(i, 4, v)
                                                if partnumber == '210-ADUX':
                                                    v = 2
                                                    w_sheet.write(i, 4, v)

                                                if partnumber == 'A9748229':
                                                    v = 2
                                                    w_sheet.write(i, 4, v)

                                            # Terminal Server Entry
                                            for i in range(2, num_rows):
                                                for j in range(0, num_cells):
                                                    partnumber = sh.cell(i, j).value
                                                    if partnumber == 'Terminal Server':
                                                        start_row = i
                                                        break
                                            start_row1 = start_row

                                            for i in range(start_row1, num_rows):
                                                partnumber = sh.cell(i, 0).value
                                                if partnumber == "":
                                                    end_row12 = i
                                                    break
                                            OOB_CORE_end_row = end_row12
                                            for i in range(start_row1, OOB_CORE_end_row):
                                                partnumber = sh.cell(i, 2).value
                                                if partnumber == 'SLC80482201S':
                                                    v = flr + 2
                                                    w_sheet.write(i, 4, v)

                                            # MGMT Leaves LAYER ENTRY
                                            for i in range(2, num_rows):
                                                for j in range(0, num_cells):
                                                    partnumber = sh.cell(i, j).value
                                                    if partnumber == 'Management Access Leaves':
                                                        start_row = i
                                                        break
                                            start_row1 = start_row

                                            for i in range(start_row1, num_rows):

                                                partnumber = sh.cell(i, 0).value
                                                if partnumber == "":
                                                    end_row12 = i
                                                    break
                                            OOB_Leaves_end_row = end_row12
                                            for i in range(start_row1, num_rows):
                                                partnumber = sh.cell(i, 2).value
                                                if partnumber == '407-BBOU-US':
                                                    v = flr * 2
                                                    w_sheet.write(i, 4, v)

                                                if partnumber == '4610-54P-O-AC-F-US' or partnumber == 'A8793201':
                                                    v = flr
                                                    w_sheet.write(i, 4, v)

                                        # -------------------------------------------------
                                        #       CODE FOR saving xl file after updating
                                        # -------------------------------------------------
                                        now = datetime.now()
                                        s = str('result{}.xls'.format(now.strftime("%c")))
                                        wb.save(s.replace(':', '_'))
                                        pyexcel.save_book_as(file_name=s.replace(':', '_'),
                                                             dest_file_name='result{}.xlsx'.format(
                                                                 now.strftime("%c")).replace(':',
                                                                                             '_'))
                                        s = 'result{}.xlsx'.format(now.strftime("%c")).replace(':', '_')

                                        # -------------------------------------------------
                                        #       Used openpyxl for editing the xlsx file
                                        # -------------------------------------------------

                                        import openpyxl
                                        from openpyxl import Workbook

                                        wb = openpyxl.load_workbook(s)
                                        sheet = wb['DETAIL-CUMULUS-BOM-SMALL-SITE']
                                        sheet1 = wb['SUMMARY-CUMULUS-BOM']

                                        # -------------------------------------------------
                                        #       CODE FOR col width(adjust the col width according to text length in it)
                                        # -------------------------------------------------
                                        dims = {}
                                        for row in sheet.rows:

                                            for cell in row:
                                                if cell.value:
                                                    dims[cell.column] = max(
                                                        (dims.get(cell.column, 0), len(str(cell.value))))
                                        for col, value in dims.items():
                                            sheet.column_dimensions[col].width = value

                                        sheet.column_dimensions['E'].width = 10
                                        sheet1.column_dimensions['A'].width = 15
                                        sheet1.column_dimensions['B'].width = 25
                                        sheet1.column_dimensions['C'].width = 25
                                        sheet1.column_dimensions['D'].width = 110
                                        sheet1.column_dimensions['E'].width = 10

                                        # -------------------------------------------------
                                        #      Cell Formatting for small site
                                        # -------------------------------------------------

                                        top_left11_cell = sheet1['A2']

                                        top_left_cell = sheet['A2']
                                        top_left1_cell = sheet['A9']
                                        top_left2_cell = sheet['A15']
                                        top_left3_cell = sheet['A22']
                                        top_left4_cell = sheet['A26']
                                        sheet.merge_cells('A2:E2')
                                        sheet.merge_cells('A9:E9')
                                        sheet.merge_cells('A15:E15')
                                        sheet.merge_cells('A22:E22')
                                        sheet.merge_cells('A26:E26')

                                        sheet1.merge_cells('A2:E2')

                                        top_left_cell.alignment = Alignment(horizontal="center", vertical="center")
                                        top_left1_cell.alignment = Alignment(horizontal="center", vertical="center")
                                        top_left2_cell.alignment = Alignment(horizontal="center", vertical="center")
                                        top_left3_cell.alignment = Alignment(horizontal="center", vertical="center")
                                        top_left4_cell.alignment = Alignment(horizontal="center", vertical="center")

                                        top_left11_cell.alignment = Alignment(horizontal="center",
                                                                              vertical="center")

                                        sheet.cell(row=2, column=1).fill = PatternFill(start_color='FFFF99',
                                                                                       end_color='FFFF99',
                                                                                       fill_type='solid')
                                        sheet.cell(row=9, column=1).fill = PatternFill(start_color='FFFF99',
                                                                                       end_color='FFFF99',
                                                                                       fill_type='solid')
                                        sheet.cell(row=15, column=1).fill = PatternFill(start_color='FFFF99',
                                                                                        end_color='FFFF99',
                                                                                        fill_type='solid')
                                        sheet.cell(row=22, column=1).fill = PatternFill(start_color='FFFF99',
                                                                                        end_color='FFFF99',
                                                                                        fill_type='solid')

                                        sheet1.cell(row=2, column=1).fill = PatternFill(start_color='FFFF99',
                                                                                        end_color='FFFF99',
                                                                                        fill_type='solid')
                                        if (flr == 1):
                                            pass
                                        else:
                                            sheet.cell(row=26, column=1).fill = PatternFill(start_color='FFFF99',
                                                                                            end_color='FFFF99',
                                                                                            fill_type='solid')

                                        # -------------------------------------------------
                                        #      Cell Formatting END
                                        # -------------------------------------------------

                                        # -------------------------------------------------
                                        #       Remove Not USED Sheet
                                        # -------------------------------------------------

                                        wb.remove_sheet(wb.get_sheet_by_name('DETAIL-CUMULUS-BOM-MEDIUM-SITE'))
                                        wb.remove_sheet(wb.get_sheet_by_name('CUMULUS Large-Site-Type'))
                                        wb.remove_sheet(wb.get_sheet_by_name('CISCO Small-Site-Type'))
                                        wb.remove_sheet(wb.get_sheet_by_name('CISCO Medium-Site-Type'))
                                        wb.remove_sheet(wb.get_sheet_by_name('CISCO Large-Site-Type'))

                                        # -------------------------------------------------
                                        #      Removed Not USED Sheet Code END
                                        # -------------------------------------------------

                                        wb.save(s)

                                        # -------------------------------------------------
                                        #       CODE FOR MAIL STARTS
                                        # -------------------------------------------------

                                    '''subject = "An email with attachment from Python"
                                        body = "This is an email with attachment sent from Python"
                                        sender_email = "pythontesting13june@gmail.com"
                                        receiver_email = "er.akash.dhand@gmail.com"
                                        password = "python@1234"

                                        # Create a multipart message and set headers
                                        message = MIMEMultipart()
                                        message["From"] = sender_email
                                        message["To"] = receiver_email
                                        message["Subject"] = subject
                                        message["Bcc"] = receiver_email  # Recommended for mass emails

                                        # Add body to email
                                        message.attach(MIMEText(body, "plain"))

                                        filename = 'result{}.xlsx'.format(now.strftime("%c")).replace(':','_')  # In same directory as script

                                        # Open PDF file in binary mode
                                        with open(filename, "rb") as attachment:
                                            # Add file as application/octet-stream
                                            # Email client can usually download this automatically as attachment
                                            part = MIMEBase("application", "octet-stream")
                                            part.set_payload(attachment.read())

                                        # Encode file in ASCII characters to send by email
                                        encoders.encode_base64(part)

                                        # Add header as key/value pair to attachment part
                                        part.add_header(
                                            "Content-Disposition",
                                            f"attachment; filename= {filename}",
                                        )

                                        # Add attachment to message and convert message to string
                                        message.attach(part)
                                        text = message.as_string()

                                        # Log in to server using secure context and send email
                                        context = ssl.create_default_context()
                                        with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=context) as server:
                                            server.login(sender_email, password)
                                            server.sendmail(sender_email, receiver_email, text)

                                        # -------------------------------------------------
                                        #       CODE FOR MAIL ENDS
                                        # -------------------------------------------------
                                    '''

                                    # -------------------------------------------------
                                    #       XL function ends
                                    # -------------------------------------------------

                                    def XL_CUMULUS_Medium_Site_Type():
                                        start_row = 0
                                        start_row111 = 0
                                        start_row1 = 0
                                        start_row2 = 0
                                        start_row3 = 0
                                        start_row4 = 0
                                        start_row5 = 0
                                        end_row = 1
                                        end_row12 = 1
                                        end_row22 = 1
                                        end_row32 = 1
                                        end_row42 = 1
                                        # -------------------------------------------------
                                        #       CODE FOR creating user access layer and core layer Dictionaries
                                        # -------------------------------------------------
                                        wb = xlrd.open_workbook('BOM Code.xls')
                                        sh = wb.sheet_by_name('DETAIL-CUMULUS-BOM-MEDIUM-SITE')
                                        num_rows = sh.nrows
                                        num_cells = sh.ncols - 1
                                        sh1 = wb.sheet_by_name('SUMMARY-CUMULUS-BOM')
                                        num_rows1 = sh1.nrows
                                        num_cells1 = sh1.ncols - 1

                                        curr_row = 0

                                        # -------------------------------------------------
                                        #       CODE FOR writing back to xl
                                        # -------------------------------------------------
                                        print("Medum xl created")
                                        rb = xlrd.open_workbook("BOM Code.xls")
                                        sh = rb.sheet_by_name('DETAIL-CUMULUS-BOM-MEDIUM-SITE')
                                        sh1 = rb.sheet_by_name('SUMMARY-CUMULUS-BOM')
                                        wb = copy(rb)
                                        w_sheet = wb.get_sheet('DETAIL-CUMULUS-BOM-MEDIUM-SITE')
                                        w_sheet1 = wb.get_sheet('SUMMARY-CUMULUS-BOM')

                                        # -------------------------------------------------
                                        #       CODE TO GET VALUES INTO CUMULUS-SUMMARY-TAB
                                        # -------------------------------------------------
                                        # SUMMARY-PAGE

                                        for i in range(2, num_rows1):
                                            partnumber = sh1.cell(i, 2).value
                                            if partnumber == '407-BBOU-US':
                                                if flr == 1:
                                                    v = leaf_list1 * 4 + 46 + flr * 4
                                                    w_sheet1.write(i, 4, v)
                                                else:
                                                    v = leaf_list1 * 4 + 46 + flr * 4 + 6
                                                    w_sheet1.write(i, 4, v)

                                            if partnumber == '407-BBOS-US':
                                                v = 80
                                                w_sheet1.write(i, 4, v)
                                            if partnumber == '210-ADUX':
                                                if flr == 1:
                                                    v = 5  # (Super Spine A and B +1 extra, SVC spine A and B )
                                                    w_sheet1.write(i, 4, v)
                                                else:
                                                    v1 = 5 + 2  # (Super Spine A and B +1 extra, SVC spine A and B, Mgmt A and B)
                                                    w_sheet1.write(i, 4, v1)
                                            if partnumber == '4610-54P-O-AC-F-US':
                                                v = leaf_list1 + 1 + flr
                                                w_sheet1.write(i, 4, v)
                                            if partnumber == 'A9748229':

                                                if flr == 1:
                                                    v1 = 2 + 2
                                                    w_sheet1.write(i, 4, v1)
                                                else:
                                                    v2 = 2 + 2 + 2
                                                    w_sheet1.write(i, 4, v2)
                                            if partnumber == 'A8793201':
                                                v = leaf_list1 + flr
                                                w_sheet1.write(i, 4, v)

                                            if partnumber == 'SLC80482201S':
                                                if flr == 1:
                                                    v = flr
                                                    w_sheet1.write(i, 4, v)
                                                else:
                                                    v = flr + 2
                                                    w_sheet1.write(i, 4, v)

                                        # -------------------------------------------------
                                        #       CODE TO GET VALUES OF ROW and insert leaf value inside qty column
                                        # -------------------------------------------------
                                        # CUMULUS SPINE ENTRY
                                        for i in range(2, num_rows):
                                            partnumber = sh.cell(i, 0).value
                                            if partnumber == "":
                                                end_row = i
                                                break
                                        Spine_end_row = end_row
                                        for i in range(2, Spine_end_row + 1):
                                            partnumber = sh.cell(i, 2).value
                                            if partnumber == '407-BBOU-US':
                                                v = leaf_list1 * 2 + 30
                                                w_sheet.write(i, 4, v)
                                            if partnumber == '407-BBOS-US':
                                                v = 20
                                                w_sheet.write(i, 4, v)
                                            if partnumber == '210-ADUX':
                                                v = 2 + 1
                                                w_sheet.write(i, 4, v)
                                            if partnumber == 'A9748229':
                                                v = 2
                                                w_sheet.write(i, 4, v)

                                        # CUMULUS SVC SPINE LAYER ENTRY
                                        for i in range(2, num_rows):
                                            for j in range(0, num_cells):
                                                partnumber = sh.cell(i, j).value
                                                if partnumber == 'SVC SPINE':
                                                    start_row = i
                                                    break
                                        start_row1 = start_row
                                        for i in range(start_row1, num_rows):
                                            partnumber = sh.cell(i, 0).value
                                            if partnumber == "":
                                                end_row12 = i
                                                break
                                        SVC_SPINE_end_row = end_row12

                                        for i in range(start_row1, SVC_SPINE_end_row + 1):
                                            partnumber = sh.cell(i, 2).value
                                            if partnumber == '407-BBOU-US':
                                                v = 14
                                                w_sheet.write(i, 4, v)
                                            if partnumber == '407-BBOS-US':
                                                v = 60
                                                w_sheet.write(i, 4, v)
                                            if partnumber == '210-ADUX':
                                                v = 2
                                                w_sheet.write(i, 4, v)

                                            if partnumber == 'A9748229':
                                                v = 2
                                                w_sheet.write(i, 4, v)

                                        # USER Access Leave ENTRY
                                        for i in range(2, num_rows):
                                            for j in range(0, num_cells):
                                                partnumber = sh.cell(i, j).value
                                                if partnumber == 'User Access Leaves':
                                                    start_row = i
                                                    break
                                        start_row2 = start_row

                                        for i in range(start_row2, num_rows):
                                            partnumber = sh.cell(i, 0).value
                                            if partnumber == "":
                                                end_row22 = i
                                                break
                                        User_Access_end_row = end_row22

                                        for i in range(start_row2, User_Access_end_row + 1):
                                            partnumber = sh.cell(i, 2).value
                                            if partnumber == '407-BBOU-US':
                                                vv = leaf_list1 * 2 + 2
                                                w_sheet.write(i, 4, vv)

                                            if partnumber == '4610-54P-O-AC-F-US':
                                                v = leaf_list1 + 1
                                                w_sheet.write(i, 4, v)
                                            if partnumber == 'A8793201':
                                                v = leaf_list1
                                                w_sheet.write(i, 4, v)

                                        # Core OOB MGMT LEAVE ENTRY
                                        if flr == 1:
                                            for i in range(2, num_rows):
                                                for j in range(0, num_cells):
                                                    partnumber = sh.cell(i, j).value
                                                    if partnumber == 'Management Core Layer':
                                                        start_row = i
                                                        break
                                            start_row1 = start_row

                                            for i in range(start_row1, num_rows):
                                                partnumber = sh.cell(i, 0).value
                                                if partnumber == "":
                                                    end_row12 = i
                                                    break
                                            OOB_CORE_end_row = end_row12
                                            for i in range(start_row1, OOB_CORE_end_row):
                                                partnumber = sh.cell(i, 2).value
                                                if partnumber == '407-BBOU-US':
                                                    v = flr * 2 + 2
                                                    w_sheet.write(i, 4, v)
                                                if partnumber == '407-BBOS-US':
                                                    v = 4
                                                    w_sheet.write(i, 4, v)
                                                if partnumber == '210-ADUX':
                                                    v = 1
                                                    w_sheet.write(i, 4, v)

                                                    b = 'EDGECORE NETWORKS'
                                                    w_sheet.write(i, 1, b)
                                                    c = '4610-54P-O-AC-F-US'
                                                    w_sheet.write(i, 2, c)
                                                    d = 'AS4610-54P 48 PORT 10/100/1000BASE-T SWITCH WITH 48 POE+ PORTS,1-8 PORT SUPPORT'
                                                    w_sheet.write(i, 3, d)
                                                if partnumber == 'A9748229':
                                                    v = 1
                                                    w_sheet.write(i, 4, v)
                                                    c1 = 'A8793201'
                                                    w_sheet.write(i, 2, c1)
                                                    d1 = 'Cumulus Linux Perpetual License, 1G, 1 Year Software Updates and Support Included'
                                                    w_sheet.write(i, 3, d1)

                                            # Terminal Server Entry
                                            for i in range(2, num_rows):
                                                for j in range(0, num_cells):
                                                    partnumber = sh.cell(i, j).value
                                                    if partnumber == 'Terminal Server':
                                                        start_row = i
                                                        break

                                            start_row1 = start_row
                                            for i in range(start_row1, num_rows):
                                                partnumber = sh.cell(i, 0).value
                                                if partnumber == "":
                                                    end_row12 = i
                                                    break
                                            OOB_CORE_end_row = end_row12
                                            for i in range(start_row1, OOB_CORE_end_row):
                                                partnumber = sh.cell(i, 2).value
                                                if partnumber == 'SLC80482201S':
                                                    v = flr
                                                    w_sheet.write(i, 4, v)
                                            # MGMT Access LAYER REMOVAL

                                            for i in range(2, num_rows):
                                                for j in range(0, num_cells):
                                                    partnumber = sh.cell(i, j).value
                                                    if partnumber == 'Management Access Leaves':
                                                        start_row = i
                                                        break
                                            start_row1 = start_row

                                            for i in range(OOB_CORE_end_row, num_rows):
                                                for j in range(num_cells + 1):
                                                    w_sheet.write(i, j, "")

                                        else:
                                            for i in range(2, num_rows):
                                                for j in range(0, num_cells):
                                                    partnumber = sh.cell(i, j).value
                                                    if partnumber == 'Management Core Layer':
                                                        start_row = i
                                                        break
                                            start_row1 = start_row

                                            for i in range(start_row1, num_rows):
                                                partnumber = sh.cell(i, 0).value
                                                if partnumber == "":
                                                    end_row12 = i
                                                    break
                                            OOB_CORE_end_row = end_row12
                                            for i in range(start_row1, OOB_CORE_end_row):
                                                partnumber = sh.cell(i, 2).value
                                                if partnumber == '407-BBOU-US':
                                                    v = flr * 2 + 6
                                                    w_sheet.write(i, 4, v)
                                                if partnumber == '407-BBOS-US':
                                                    v = 4
                                                    w_sheet.write(i, 4, v)
                                                if partnumber == '210-ADUX':
                                                    v = 2
                                                    w_sheet.write(i, 4, v)

                                                if partnumber == 'A9748229':
                                                    v = 2
                                                    w_sheet.write(i, 4, v)
                                            # Terminal Server Entry
                                            for i in range(2, num_rows):
                                                for j in range(0, num_cells):
                                                    partnumber = sh.cell(i, j).value
                                                    if partnumber == 'Terminal Server':
                                                        start_row = i
                                                        break

                                            start_row1 = start_row
                                            for i in range(start_row1, num_rows):
                                                partnumber = sh.cell(i, 0).value
                                                if partnumber == "":
                                                    end_row12 = i
                                                    break
                                            OOB_CORE_end_row = end_row12
                                            for i in range(start_row1, OOB_CORE_end_row):
                                                partnumber = sh.cell(i, 2).value
                                                if partnumber == 'SLC80482201S':
                                                    v = flr + 2
                                                    w_sheet.write(i, 4, v)
                                            # MGMT Leaves LAYER ENTRY
                                            for i in range(2, num_rows):
                                                for j in range(0, num_cells):
                                                    partnumber = sh.cell(i, j).value
                                                    if partnumber == 'Management Access Leaves':
                                                        start_row = i
                                                        break
                                            start_row1 = start_row

                                            for i in range(start_row1, num_rows):

                                                partnumber = sh.cell(i, 0).value
                                                if partnumber == "":
                                                    end_row12 = i
                                                    break
                                            OOB_Leaves_end_row = end_row12
                                            for i in range(start_row1, num_rows):
                                                partnumber = sh.cell(i, 2).value
                                                if partnumber == '407-BBOU-US':
                                                    v = flr * 2
                                                    w_sheet.write(i, 4, v)

                                                if partnumber == '4610-54P-O-AC-F-US' or partnumber == 'A8793201':
                                                    v = flr
                                                    w_sheet.write(i, 4, v)
                                        # -------------------------------------------------
                                        #       CODE FOR saving xl file after updating
                                        # -------------------------------------------------
                                        now = datetime.now()
                                        s = str('result{}.xls'.format(now.strftime("%c")))
                                        wb.save(s.replace(':', '_'))
                                        pyexcel.save_book_as(file_name=s.replace(':', '_'),
                                                             dest_file_name='result{}.xlsx'.format(
                                                                 now.strftime("%c")).replace(':',
                                                                                             '_'))
                                        s = 'result{}.xlsx'.format(now.strftime("%c")).replace(':', '_')
                                        # -------------------------------------------------
                                        #       Used openpyxl for editing the xlsx file
                                        # -------------------------------------------------
                                        import openpyxl
                                        from openpyxl.styles import Font
                                        wb = openpyxl.load_workbook(s)
                                        sheet = wb['DETAIL-CUMULUS-BOM-MEDIUM-SITE']
                                        sheet1 = wb['SUMMARY-CUMULUS-BOM']
                                        # -------------------------------------------------
                                        #       CODE FOR col width(adjust the col width according to text length in it)
                                        # -------------------------------------------------
                                        dims = {}
                                        for row in sheet.rows:
                                            for cell in row:
                                                if cell.value:
                                                    dims[cell.column] = max(
                                                        (dims.get(cell.column, 0), len(str(cell.value))))
                                        for col, value in dims.items():
                                            sheet.column_dimensions[col].width = value

                                        sheet.column_dimensions['E'].width = 10
                                        sheet1.column_dimensions['A'].width = 15
                                        sheet1.column_dimensions['B'].width = 25
                                        sheet1.column_dimensions['C'].width = 25
                                        sheet1.column_dimensions['D'].width = 110
                                        sheet1.column_dimensions['E'].width = 10

                                        # -------------------------------------------------
                                        #      Cell Formatting
                                        # -------------------------------------------------

                                        top_left11_cell = sheet1['A2']

                                        top_left_cell = sheet['A2']
                                        top_left1_cell = sheet['A9']
                                        top_left2_cell = sheet['A16']
                                        top_left3_cell = sheet['A22']
                                        top_left4_cell = sheet['A28']
                                        top_left5_cell = sheet['A32']
                                        sheet.merge_cells('A2:E2')
                                        sheet.merge_cells('A9:E9')
                                        sheet.merge_cells('A16:E16')
                                        sheet.merge_cells('A22:E22')
                                        sheet.merge_cells('A28:E28')
                                        sheet.merge_cells('A32:E32')

                                        sheet1.merge_cells('A2:E2')

                                        top_left_cell.alignment = Alignment(horizontal="center", vertical="center")
                                        top_left1_cell.alignment = Alignment(horizontal="center", vertical="center")
                                        top_left2_cell.alignment = Alignment(horizontal="center", vertical="center")
                                        top_left3_cell.alignment = Alignment(horizontal="center", vertical="center")
                                        top_left4_cell.alignment = Alignment(horizontal="center", vertical="center")
                                        top_left5_cell.alignment = Alignment(horizontal="center", vertical="center")

                                        top_left11_cell.alignment = Alignment(horizontal="center",
                                                                              vertical="center")

                                        sheet.cell(row=2, column=1).fill = PatternFill(start_color='FFFF99',
                                                                                       end_color='FFFF99',
                                                                                       fill_type='solid')
                                        sheet.cell(row=9, column=1).fill = PatternFill(start_color='FFFF99',
                                                                                       end_color='FFFF99',
                                                                                       fill_type='solid')
                                        sheet.cell(row=16, column=1).fill = PatternFill(start_color='FFFF99',
                                                                                        end_color='FFFF99',
                                                                                        fill_type='solid')

                                        sheet.cell(row=22, column=1).fill = PatternFill(start_color='FFFF99',
                                                                                        end_color='FFFF99',
                                                                                        fill_type='solid')

                                        sheet.cell(row=28, column=1).fill = PatternFill(start_color='FFFF99',
                                                                                        end_color='FFFF99',
                                                                                        fill_type='solid')

                                        sheet1.cell(row=2, column=1).fill = PatternFill(start_color='FFFF99',
                                                                                        end_color='FFFF99',
                                                                                        fill_type='solid')
                                        if flr == 1:
                                            pass
                                        else:
                                            sheet.cell(row=32, column=1).fill = PatternFill(start_color='FFFF99',
                                                                                            end_color='FFFF99',
                                                                                            fill_type='solid')

                                        # -------------------------------------------------
                                        #      Cell Formatting END
                                        # -------------------------------------------------

                                        # -------------------------------------------------
                                        #       Remove Not USED Sheet
                                        # -------------------------------------------------
                                        wb.remove_sheet(wb.get_sheet_by_name('DETAIL-CUMULUS-BOM-SMALL-SITE'))
                                        wb.remove_sheet(wb.get_sheet_by_name('CUMULUS Large-Site-Type'))
                                        wb.remove_sheet(wb.get_sheet_by_name('CISCO Small-Site-Type'))
                                        wb.remove_sheet(wb.get_sheet_by_name('CISCO Medium-Site-Type'))
                                        wb.remove_sheet(wb.get_sheet_by_name('CISCO Large-Site-Type'))

                                        # -------------------------------------------------
                                        #      Removed Not USED Sheet Code END
                                        # -------------------------------------------------

                                        wb.save(s)

                                        # -------------------------------------------------
                                        #       CODE FOR MAIL STARTS
                                        # -------------------------------------------------
                                        '''subject = "An email with attachment from Python"
                                        body = "This is an email with attachment sent from Python"
                                        sender_email = "pythontesting13june@gmail.com"
                                        receiver_email = "er.akash.dhand@gmail.com"
                                        password = "python@1234"

                                        # Create a multipart message and set headers
                                        message = MIMEMultipart()
                                        message["From"] = sender_email
                                        message["To"] = receiver_email
                                        message["Subject"] = subject
                                        message["Bcc"] = receiver_email  # Recommended for mass emails

                                        # Add body to email
                                        message.attach(MIMEText(body, "plain"))

                                        filename = 'result{}.xlsx'.format(now.strftime("%c")).replace(':','_')  # In same directory as script

                                        # Open PDF file in binary mode
                                        with open(filename, "rb") as attachment:
                                            # Add file as application/octet-stream
                                            # Email client can usually download this automatically as attachment
                                            part = MIMEBase("application", "octet-stream")
                                            part.set_payload(attachment.read())

                                        # Encode file in ASCII characters to send by email
                                        encoders.encode_base64(part)

                                        # Add header as key/value pair to attachment part
                                        part.add_header(
                                            "Content-Disposition",
                                            f"attachment; filename= {filename}",
                                        )

                                        # Add attachment to message and convert message to string
                                        message.attach(part)
                                        text = message.as_string()

                                        # Log in to server using secure context and send email
                                        context = ssl.create_default_context()
                                        with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=context) as server:
                                            server.login(sender_email, password)
                                            server.sendmail(sender_email, receiver_email, text)

                                        # -------------------------------------------------
                                        #       CODE FOR MAIL ENDS
                                        # -------------------------------------------------
                                    '''

                                    # -------------------------------------------------
                                    #       XL function ends
                                    # -------------------------------------------------
                                    # ---------------------------------------------------------
                                    #           Code for large excel site(Starts)
                                    # ---------------------------------------------------------
                                    def XL_CUMULUS_Large_Site_Type():
                                        start_row = 0
                                        end_row = 1
                                        end_row12 = 0
                                        start_row_dist = 0
                                        start_row_access = 0
                                        start_row_service = 0
                                        end_row_dist = 0

                                        # -------------------------------------------------
                                        #       CODE FOR creating user access layer and core layer Dictionaries
                                        # -------------------------------------------------

                                        wb = xlrd.open_workbook('BOM Code.xls')
                                        sh = wb.sheet_by_name('CUMULUS Large-Site-Type')
                                        num_rows = sh.nrows
                                        num_cells = sh.ncols - 1

                                        sh1 = wb.sheet_by_name('SUMMARY-CUMULUS-BOM')
                                        num_rows1 = sh1.nrows
                                        num_cells1 = sh1.ncols - 1

                                        curr_row = 0
                                        # -------------------------------------------------
                                        #       CODE FOR writing back to xl
                                        # -------------------------------------------------
                                        print("xl created")
                                        rb = xlrd.open_workbook("BOM Code.xls")
                                        sh = rb.sheet_by_name('CUMULUS Large-Site-Type')
                                        sh1 = rb.sheet_by_name('SUMMARY-CUMULUS-BOM')
                                        wb = copy(rb)
                                        w_sheet = wb.get_sheet('CUMULUS Large-Site-Type')
                                        w_sheet1 = wb.get_sheet('SUMMARY-CUMULUS-BOM')

                                        # -------------------------------------------------
                                        #       CODE TO GET VALUES INTO CUMULUS-SUMMARY-TAB
                                        # -------------------------------------------------
                                        # SUMMARY-PAGE

                                        for i in range(2, num_rows1):
                                            partnumber = sh1.cell(i, 2).value
                                            if partnumber == '407-BBOU-US':
                                                if flr == 1:
                                                    if leaf_list1 >= 1 and leaf_list1 <= 40:
                                                        v = 2 * 2 + leaf_list1 * 4 + 46 + 14 + flr * 2  # (2 spine)*2+ Leaf*4+ 30(for Super spine)+16(for Spine) Extra)+ 14 for SVC spine
                                                        w_sheet1.write(i, 4, v)
                                                    if leaf_list1 >= 41 and leaf_list1 <= 80:
                                                        v = 4 * 2 + leaf_list1 * 4 + 62 + 14 + flr * 2  # (4 spine)*2+ Leaf*4+ 30(for Super spine)+32(for Spine) Extra)+14 for SVC spine
                                                        w_sheet1.write(i, 4, v)
                                                    if leaf_list1 >= 81 and leaf_list1 <= 120:
                                                        v = 6 * 2 + leaf_list1 * 4 + 78 + 14 + flr * 2  # (6 spine)*2+ Leaf*4+ 30(for Super spine)+48(for Spine) Extra)+14 for SVC spine
                                                        w_sheet1.write(i, 4, v)
                                                    if leaf_list1 >= 121 and leaf_list1 <= 160:
                                                        v = 8 * 2 + leaf_list1 * 4 + 94 + 14 + flr * 2  # (8 spine)*2+ Leaf*4+ 30(for Super spine)+64(for Spine) Extra)+14 for SVC spine
                                                        w_sheet1.write(i, 4, v)
                                                    if leaf_list1 >= 161 and leaf_list1 <= 200:
                                                        v = 10 * 2 + leaf_list1 * 4 + 110 + 14 + flr * 2  # (10 spine)*2+ Leaf*4+ 30(for Super spine)+80(for Spine) Extra)+14 for SVC spine
                                                        w_sheet1.write(i, 4, v)
                                                    if leaf_list1 >= 201 and leaf_list1 <= 240:
                                                        v = 12 * 2 + leaf_list1 * 4 + 126 + 14 + flr * 2  # (12 spine)*2+ Leaf*4+ 30(for Super spine)+96(for Spine) Extra)+14 for SVC spine
                                                        w_sheet1.write(i, 4, v)
                                                    if leaf_list1 >= 241 and leaf_list1 <= 280:
                                                        v = 14 * 2 + leaf_list1 * 4 + 142 + 14 + flr * 2  # (14 spine)*2+ Leaf*4+ 30(for Super spine)+112(for Spine) Extra)+14 for SVC spine
                                                        w_sheet1.write(i, 4, v)
                                                    if leaf_list1 >= 281 and leaf_list1 <= 320:
                                                        v = 16 * 2 + leaf_list1 * 4 + 158 + 14 + flr * 2  # (16 spine)*2+ Leaf*4+ 30(for Super spine)+128(for Spine) Extra)+14 for SVC spine
                                                        w_sheet1.write(i, 4, v)
                                                    if leaf_list1 >= 321 and leaf_list1 <= 360:
                                                        v = 18 * 2 + leaf_list1 * 4 + 174 + 14 + flr * 2  # (18 spine)*2+ Leaf*4+ 30(for Super spine)+144(for Spine) Extra)+14 for SVC spine
                                                        w_sheet1.write(i, 4, v)
                                                    if leaf_list1 >= 361 and leaf_list1 <= 400:
                                                        v = 20 * 2 + leaf_list1 * 4 + 190 + 14 + flr * 2  # (20 spine)*2+ Leaf*4+ 30(for Super spine)+160(for Spine) Extra)+14 for SVC spine
                                                        w_sheet1.write(i, 4, v)
                                                    if leaf_list1 >= 401 and leaf_list1 <= 440:
                                                        v = 22 * 2 + leaf_list1 * 4 + 206 + 14 + flr * 2  # (20 spine)*2+ Leaf*4+ 30(for Super spine)+176(for Spine) Extra)+14 for SVC spine
                                                        w_sheet1.write(i, 4, v)
                                                else:
                                                    if leaf_list1 >= 1 and leaf_list1 <= 40:
                                                        v = 2 * 2 + leaf_list1 * 4 + 46 + 14 + flr * 4 + 6  # (2 spine)*2+ Leaf*4+ 30(for Super spine)+16(for Spine) Extra)+ 14 for SVC spine
                                                        w_sheet1.write(i, 4, v)
                                                    if leaf_list1 >= 41 and leaf_list1 <= 80:
                                                        v = 4 * 2 + leaf_list1 * 4 + 62 + 14 + flr * 4 + 6  # (4 spine)*2+ Leaf*4+ 30(for Super spine)+32(for Spine) Extra)+14 for SVC spine
                                                        w_sheet1.write(i, 4, v)
                                                    if leaf_list1 >= 81 and leaf_list1 <= 120:
                                                        v = 6 * 2 + leaf_list1 * 4 + 78 + 14 + flr * 4 + 6  # (6 spine)*2+ Leaf*4+ 30(for Super spine)+48(for Spine) Extra)+14 for SVC spine
                                                        w_sheet1.write(i, 4, v)
                                                    if leaf_list1 >= 121 and leaf_list1 <= 160:
                                                        v = 8 * 2 + leaf_list1 * 4 + 94 + 14 + flr * 4 + 6  # (8 spine)*2+ Leaf*4+ 30(for Super spine)+64(for Spine) Extra)+14 for SVC spine
                                                        w_sheet1.write(i, 4, v)
                                                    if leaf_list1 >= 161 and leaf_list1 <= 200:
                                                        v = 10 * 2 + leaf_list1 * 4 + 110 + 14 + flr * 4 + 6  # (10 spine)*2+ Leaf*4+ 30(for Super spine)+80(for Spine) Extra)+14 for SVC spine
                                                        w_sheet1.write(i, 4, v)
                                                    if leaf_list1 >= 201 and leaf_list1 <= 240:
                                                        v = 12 * 2 + leaf_list1 * 4 + 126 + 14 + flr * 4 + 6  # (12 spine)*2+ Leaf*4+ 30(for Super spine)+96(for Spine) Extra)+14 for SVC spine
                                                        w_sheet1.write(i, 4, v)
                                                    if leaf_list1 >= 241 and leaf_list1 <= 280:
                                                        v = 14 * 2 + leaf_list1 * 4 + 142 + 14 + flr * 4 + 6  # (14 spine)*2+ Leaf*4+ 30(for Super spine)+112(for Spine) Extra)+14 for SVC spine
                                                        w_sheet1.write(i, 4, v)
                                                    if leaf_list1 >= 281 and leaf_list1 <= 320:
                                                        v = 16 * 2 + leaf_list1 * 4 + 158 + 14 + flr * 4 + 6  # (16 spine)*2+ Leaf*4+ 30(for Super spine)+128(for Spine) Extra)+14 for SVC spine
                                                        w_sheet1.write(i, 4, v)
                                                    if leaf_list1 >= 321 and leaf_list1 <= 360:
                                                        v = 18 * 2 + leaf_list1 * 4 + 174 + 14 + flr * 4 + 6  # (18 spine)*2+ Leaf*4+ 30(for Super spine)+144(for Spine) Extra)+14 for SVC spine
                                                        w_sheet1.write(i, 4, v)
                                                    if leaf_list1 >= 361 and leaf_list1 <= 400:
                                                        v = 20 * 2 + leaf_list1 * 4 + 190 + 14 + flr * 4 + 6  # (20 spine)*2+ Leaf*4+ 30(for Super spine)+160(for Spine) Extra)+14 for SVC spine
                                                        w_sheet1.write(i, 4, v)
                                                    if leaf_list1 >= 401 and leaf_list1 <= 440:
                                                        v = 22 * 2 + leaf_list1 * 4 + 206 + 14 + flr * 4 + 6  # (20 spine)*2+ Leaf*4+ 30(for Super spine)+176(for Spine) Extra)+14 for SVC spine
                                                        w_sheet1.write(i, 4, v)

                                            if partnumber == '407-BBOS-US':
                                                if leaf_list1 >= 1 and leaf_list1 <= 40:
                                                    v = 4 + 10 + 80 + 4  # 10 For Super Spine +4 for Spine + 80 for SVC spine + 4 for MGT
                                                    w_sheet1.write(i, 4, v)
                                                if leaf_list1 >= 41 and leaf_list1 <= 80:
                                                    v = 8 + 10 + 80 + 4  # 10 For Super Spine +8 for Spine + 80 for SVC spine + 4 for MGT
                                                    w_sheet1.write(i, 4, v)
                                                if leaf_list1 >= 81 and sum(leaf_list) <= 120:
                                                    v = 12 + 10 + 80 + 4  # 10 For Super Spine +12 for Spine + 80 for SVC spine + 4 for MGT
                                                    w_sheet1.write(i, 4, v)
                                                if leaf_list1 >= 121 and leaf_list1 <= 160:
                                                    v = 16 + 10 + 80 + 4  # 10 For Super Spine +16 for Spine + 80 for SVC spine + 4 for MGT
                                                    w_sheet1.write(i, 4, v)
                                                if leaf_list1 >= 161 and leaf_list1 <= 200:
                                                    v = 20 + 10 + 80 + 4  # 10 For Super Spine +20 for Spine + 80 for SVC spine + 4 for MGT
                                                    w_sheet1.write(i, 4, v)
                                                if leaf_list1 >= 201 and leaf_list1 <= 240:
                                                    v = 24 + 10 + 80 + 4  # 10 For Super Spine +24 for Spine + 80 for SVC spine + 4 for MGT
                                                    w_sheet1.write(i, 4, v)
                                                if leaf_list1 >= 241 and leaf_list1 <= 280:
                                                    v = 28 + 10 + 80 + 4  # 10 For Super Spine +28 for Spine + 80 for SVC spine + 4 for MGT
                                                    w_sheet1.write(i, 4, v)
                                                if leaf_list1 >= 281 and leaf_list1 <= 320:
                                                    v = 32 + 10 + 80 + 4  # 10 For Super Spine +32 for Spine + 80 for SVC spine + 4 for MGT
                                                    w_sheet1.write(i, 4, v)
                                                if leaf_list1 >= 321 and leaf_list1 <= 360:
                                                    v = 36 + 10 + 80 + 4  # 10 For Super Spine +36 for Spine + 80 for SVC spine + 4 for MGT
                                                    w_sheet1.write(i, 4, v)
                                                if leaf_list1 >= 361 and leaf_list1 <= 400:
                                                    v = 40 + 10 + 80 + 4  # 10 For Super Spine +40 for Spine + 80 for SVC spine + 4 for MGT
                                                    w_sheet1.write(i, 4, v)
                                                if leaf_list1 >= 401 and leaf_list1 <= 440:
                                                    v = 44 + 10 + 80 + 4  # 10 For Super Spine +44 for Spine + 80 for SVC spine + 4 for MGT
                                                    w_sheet1.write(i, 4, v)

                                            if partnumber == '210-ADUX':
                                                if flr == 1:
                                                    if leaf_list1 >= 1 and leaf_list1 <= 40:
                                                        v = 2 + 2 + 1 + 2  # (2 For Super Spine , Spine A and Spine B +1, 2 SVC spine)
                                                        w_sheet1.write(i, 4, v)
                                                    if leaf_list1 >= 41 and leaf_list1 <= 80:
                                                        v = 2 + 4 + 1 + 2  # (2 For Super Spine ,Spine A,b,c,d +1, 2 SVC spine)
                                                        w_sheet1.write(i, 4, v)
                                                    if leaf_list1 >= 81 and leaf_list1 <= 120:
                                                        v = 2 + 6 + 1 + 2  # (2 For Super Spine ,Spine A,b,c,d,e,f +1, 2 SVC spine)
                                                        w_sheet1.write(i, 4, v)
                                                    if leaf_list1 >= 121 and leaf_list1 <= 160:
                                                        v = 2 + 8 + 1 + 2  # (2 For Super Spine ,Spine A,b,c,d,e,f,g,h +1, 2 SVC spine)
                                                        w_sheet1.write(i, 4, v)
                                                    if leaf_list1 >= 161 and leaf_list1 <= 200:
                                                        v = 2 + 10 + 1 + 2  # (2 For Super Spine ,Spine A,b,c,d,e,f,g,h,i,j +1, 2 SVC spine)
                                                        w_sheet1.write(i, 4, v)
                                                    if leaf_list1 >= 201 and leaf_list1 <= 240:
                                                        v = 2 + 12 + 1 + 2  # (2 For Super Spine ,Spine A,b,c,d,e,f,g,h,i,j,k,l +1, 2 SVC spine)
                                                        w_sheet1.write(i, 4, v)
                                                    if leaf_list1 >= 241 and leaf_list1 <= 280:
                                                        v = 2 + 14 + 1 + 2  # (2 For Super Spine ,Spine A,b,c,d,e,f,g,h,i,j,k,l,m,n +1, 2 SVC spine)
                                                        w_sheet1.write(i, 4, v)
                                                    if leaf_list1 >= 281 and leaf_list1 <= 320:
                                                        v = 2 + 16 + 1 + 2  # (2 For Super Spine ,Spine A,b,c,d,e,f,g,h,i,j,k,l,m,n,o,p +1, 2 SVC spine)
                                                        w_sheet1.write(i, 4, v)
                                                    if leaf_list1 >= 321 and leaf_list1 <= 360:
                                                        v = 2 + 18 + 1 + 2  # (2 For Super Spine ,Spine A,b,c,d,e,f,g,h,i,j,k,l,m,n,o,p,q,r +1, 2 SVC spine)
                                                        w_sheet1.write(i, 4, v)
                                                    if leaf_list1 >= 361 and leaf_list1 <= 400:
                                                        v = 2 + 20 + 1 + 2  # (2 For Super Spine ,Spine A,b,c,d,e,f,g,h,i,j,k,l,m,n,o,p,q,r,s,t +1, 2 SVC spine)
                                                        w_sheet1.write(i, 4, v)
                                                    if leaf_list1 >= 401 and leaf_list1 <= 440:
                                                        v = 2 + 22 + 1 + 2  # (2 For Super Spine ,Spine A,b,c,d,e,f,g,h,i,j,k,l,m,n,o,p,q,r,s,t,u,v +1, 2 SVC spine)
                                                        w_sheet1.write(i, 4, v)

                                                else:
                                                    if leaf_list1 >= 1 and leaf_list1 <= 40:
                                                        v = 2 + 2 + 1 + 2 + 2  # (2 For Super Spine , Spine A and Spine B +1, 2 SVC spine+2 MGT Spine)
                                                        w_sheet1.write(i, 4, v)
                                                    if leaf_list1 >= 41 and leaf_list1 <= 80:
                                                        v = 2 + 4 + 1 + 2 + 2  # (2 For Super Spine ,Spine A,b,c,d +1, 2 SVC spine2 MGT Spine)
                                                        w_sheet1.write(i, 4, v)
                                                    if leaf_list1 >= 81 and leaf_list1 <= 120:
                                                        v = 2 + 6 + 1 + 2 + 2  # (2 For Super Spine ,Spine A,b,c,d,e,f +1, 2 SVC spine2 MGT Spine)
                                                        w_sheet1.write(i, 4, v)
                                                    if leaf_list1 >= 121 and leaf_list1 <= 160:
                                                        v = 2 + 8 + 1 + 2 + 2  # (2 For Super Spine ,Spine A,b,c,d,e,f,g,h +1, 2 SVC spine2 MGT Spine)
                                                        w_sheet1.write(i, 4, v)
                                                    if leaf_list1 >= 161 and leaf_list1 <= 200:
                                                        v = 2 + 10 + 1 + 2 + 2  # (2 For Super Spine ,Spine A,b,c,d,e,f,g,h,i,j +1, 2 SVC spine2 MGT Spine)
                                                        w_sheet1.write(i, 4, v)
                                                    if leaf_list1 >= 201 and leaf_list1 <= 240:
                                                        v = 2 + 12 + 1 + 2 + 2  # (2 For Super Spine ,Spine A,b,c,d,e,f,g,h,i,j,k,l +1, 2 SVC spine2 MGT Spine)
                                                        w_sheet1.write(i, 4, v)
                                                    if leaf_list1 >= 241 and leaf_list1 <= 280:
                                                        v = 2 + 14 + 1 + 2 + 2  # (2 For Super Spine ,Spine A,b,c,d,e,f,g,h,i,j,k,l,m,n +1, 2 SVC spine2 MGT Spine)
                                                        w_sheet1.write(i, 4, v)
                                                    if leaf_list1 >= 281 and leaf_list1 <= 320:
                                                        v = 2 + 16 + 1 + 2 + 2  # (2 For Super Spine ,Spine A,b,c,d,e,f,g,h,i,j,k,l,m,n,o,p +1, 2 SVC spine2 MGT Spine)
                                                        w_sheet1.write(i, 4, v)
                                                    if leaf_list1 >= 321 and leaf_list1 <= 360:
                                                        v = 2 + 18 + 1 + 2 + 2  # (2 For Super Spine ,Spine A,b,c,d,e,f,g,h,i,j,k,l,m,n,o,p,q,r +1, 2 SVC spine2 MGT Spine)
                                                        w_sheet1.write(i, 4, v)
                                                    if leaf_list1 >= 361 and leaf_list1 <= 400:
                                                        v = 2 + 20 + 1 + 2 + 2  # (2 For Super Spine ,Spine A,b,c,d,e,f,g,h,i,j,k,l,m,n,o,p,q,r,s,t +1, 2 SVC spine2 MGT Spine)
                                                        w_sheet1.write(i, 4, v)
                                                    if leaf_list1 >= 401 and leaf_list1 <= 440:
                                                        v = 2 + 22 + 1 + 2 + 2  # (2 For Super Spine ,Spine A,b,c,d,e,f,g,h,i,j,k,l,m,n,o,p,q,r,s,t,u,v +1, 2 SVC spine2 MGT Spine)
                                                        w_sheet1.write(i, 4, v)
                                            if partnumber == '4610-54P-O-AC-F-US':
                                                v = leaf_list1 + 1 + flr
                                                w_sheet1.write(i, 4, v)
                                            if partnumber == 'A9748229':
                                                if flr == 1:
                                                    if leaf_list1 >= 1 and leaf_list1 <= 40:
                                                        v = 2 + 2 + 2  # (2 For Super Spine , Spine A and Spine B +1, 2 SVC spine)
                                                        w_sheet1.write(i, 4, v)
                                                    if leaf_list1 >= 41 and leaf_list1 <= 80:
                                                        v = 2 + 4 + 2  # (2 For Super Spine ,Spine A,b,c,d +1, 2 SVC spine)
                                                        w_sheet1.write(i, 4, v)
                                                    if leaf_list1 >= 81 and leaf_list1 <= 120:
                                                        v = 2 + 6 + 2  # (2 For Super Spine ,Spine A,b,c,d,e,f +1, 2 SVC spine)
                                                        w_sheet1.write(i, 4, v)
                                                    if leaf_list1 >= 121 and leaf_list1 <= 160:
                                                        v = 2 + 8 + 2  # (2 For Super Spine ,Spine A,b,c,d,e,f,g,h +1, 2 SVC spine)
                                                        w_sheet1.write(i, 4, v)
                                                    if leaf_list1 >= 161 and leaf_list1 <= 200:
                                                        v = 2 + 10 + 2  # (2 For Super Spine ,Spine A,b,c,d,e,f,g,h,i,j +1, 2 SVC spine)
                                                        w_sheet1.write(i, 4, v)
                                                    if leaf_list1 >= 201 and leaf_list1 <= 240:
                                                        v = 2 + 12 + 2  # (2 For Super Spine ,Spine A,b,c,d,e,f,g,h,i,j,k,l +1, 2 SVC spine)
                                                        w_sheet1.write(i, 4, v)
                                                    if leaf_list1 >= 241 and leaf_list1 <= 280:
                                                        v = 2 + 14 + 2  # (2 For Super Spine ,Spine A,b,c,d,e,f,g,h,i,j,k,l,m,n +1, 2 SVC spine)
                                                        w_sheet1.write(i, 4, v)
                                                    if leaf_list1 >= 281 and leaf_list1 <= 320:
                                                        v = 2 + 16 + 2  # (2 For Super Spine ,Spine A,b,c,d,e,f,g,h,i,j,k,l,m,n,o,p +1, 2 SVC spine)
                                                        w_sheet1.write(i, 4, v)
                                                    if leaf_list1 >= 321 and leaf_list1 <= 360:
                                                        v = 2 + 18 + 2  # (2 For Super Spine ,Spine A,b,c,d,e,f,g,h,i,j,k,l,m,n,o,p,q,r +1, 2 SVC spine)
                                                        w_sheet1.write(i, 4, v)
                                                    if leaf_list1 >= 361 and leaf_list1 <= 400:
                                                        v = 2 + 20 + 2  # (2 For Super Spine ,Spine A,b,c,d,e,f,g,h,i,j,k,l,m,n,o,p,q,r,s,t +1, 2 SVC spine)
                                                        w_sheet1.write(i, 4, v)
                                                    if leaf_list1 >= 401 and leaf_list1 <= 440:
                                                        v = 2 + 22 + 2  # (2 For Super Spine ,Spine A,b,c,d,e,f,g,h,i,j,k,l,m,n,o,p,q,r,s,t,u,v +1, 2 SVC spine)
                                                        w_sheet1.write(i, 4, v)

                                                else:
                                                    if leaf_list1 >= 1 and leaf_list1 <= 40:
                                                        v = 2 + 2 + 2 + 2  # (2 For Super Spine , Spine A and Spine B, 2 SVC spine+2 MGT Spine)
                                                        w_sheet1.write(i, 4, v)
                                                    if leaf_list1 >= 41 and leaf_list1 <= 80:
                                                        v = 2 + 4 + 2 + 2  # (2 For Super Spine ,Spine A,b,c,d, 2 SVC spine2 MGT Spine)
                                                        w_sheet1.write(i, 4, v)
                                                    if leaf_list1 >= 81 and leaf_list1 <= 120:
                                                        v = 2 + 6 + 2 + 2  # (2 For Super Spine ,Spine A,b,c,d,e,f, 2 SVC spine2 MGT Spine)
                                                        w_sheet1.write(i, 4, v)
                                                    if leaf_list1 >= 121 and leaf_list1 <= 160:
                                                        v = 2 + 8 + 2 + 2  # (2 For Super Spine ,Spine A,b,c,d,e,f,g,h, 2 SVC spine2 MGT Spine)
                                                        w_sheet1.write(i, 4, v)
                                                    if leaf_list1 >= 161 and leaf_list1 <= 200:
                                                        v = 2 + 10 + 2 + 2  # (2 For Super Spine ,Spine A,b,c,d,e,f,g,h,i,j, 2 SVC spine2 MGT Spine)
                                                        w_sheet1.write(i, 4, v)
                                                    if leaf_list1 >= 201 and leaf_list1 <= 240:
                                                        v = 2 + 12 + 2 + 2  # (2 For Super Spine ,Spine A,b,c,d,e,f,g,h,i,j,k,l, 2 SVC spine2 MGT Spine)
                                                        w_sheet1.write(i, 4, v)
                                                    if leaf_list1 >= 241 and leaf_list1 <= 280:
                                                        v = 2 + 14 + 2 + 2  # (2 For Super Spine ,Spine A,b,c,d,e,f,g,h,i,j,k,l,m,n, 2 SVC spine2 MGT Spine)
                                                        w_sheet1.write(i, 4, v)
                                                    if leaf_list1 >= 281 and leaf_list1 <= 320:
                                                        v = 2 + 16 + 2 + 2  # (2 For Super Spine ,Spine A,b,c,d,e,f,g,h,i,j,k,l,m,n,o,p, 2 SVC spine2 MGT Spine)
                                                        w_sheet1.write(i, 4, v)
                                                    if leaf_list1 >= 321 and leaf_list1 <= 360:
                                                        v = 2 + 18 + 2 + 2  # (2 For Super Spine ,Spine A,b,c,d,e,f,g,h,i,j,k,l,m,n,o,p,q,r, 2 SVC spine2 MGT Spine)
                                                        w_sheet1.write(i, 4, v)
                                                    if leaf_list1 >= 361 and leaf_list1 <= 400:
                                                        v = 2 + 20 + 2 + 2  # (2 For Super Spine ,Spine A,b,c,d,e,f,g,h,i,j,k,l,m,n,o,p,q,r,s,t, 2 SVC spine2 MGT Spine)
                                                        w_sheet1.write(i, 4, v)
                                                    if leaf_list1 >= 401 and leaf_list1 <= 440:
                                                        v = 2 + 22 + 2 + 2  # (2 For Super Spine ,Spine A,b,c,d,e,f,g,h,i,j,k,l,m,n,o,p,q,r,s,t,u,v, 2 SVC spine2 MGT Spine)
                                                        w_sheet1.write(i, 4, v)
                                            if partnumber == 'A8793201':
                                                v = leaf_list1 + flr
                                                w_sheet1.write(i, 4, v)
                                            if partnumber == 'SLC80482201S':
                                                if flr == 1:
                                                    v = flr
                                                    w_sheet1.write(i, 4, v)
                                                else:
                                                    v = flr + 2
                                                    w_sheet1.write(i, 4, v)
                                        # ------------------------------------------------------------------------
                                        #       CODE TO GET VALUES OF ROW and insert leaf value inside qty column
                                        # -----------------------------------------------------------------------
                                        # CORE LAYER ENTRY
                                        for i in range(2, num_rows):
                                            partnumber = sh.cell(i, 0).value
                                            if partnumber == "":
                                                end_row = i
                                                break
                                        Spine_end_row = end_row
                                        for i in range(2, Spine_end_row + 1):
                                            partnumber = sh.cell(i, 2).value
                                            if partnumber == '407-BBOU-US':
                                                if leaf_list1 >= 1 and leaf_list1 <= 40:
                                                    v = 2 * 2 + 30  # (Spine*2+30 Extra)
                                                    w_sheet.write(i, 4, v)
                                                if leaf_list1 >= 41 and leaf_list1 <= 80:
                                                    v = 4 * 2 + 30  # (Spine*2+30 Extra)
                                                    w_sheet.write(i, 4, v)
                                                if leaf_list1 >= 81 and leaf_list1 <= 120:
                                                    v = 6 * 2 + 30  # (Spine*2+30 Extra)
                                                    w_sheet.write(i, 4, v)
                                                if leaf_list1 >= 121 and leaf_list1 <= 160:
                                                    v = 8 * 2 + 30  # (Spine*2+30 Extra)
                                                    w_sheet.write(i, 4, v)
                                                if leaf_list1 >= 161 and leaf_list1 <= 200:
                                                    v = 10 * 2 + 30  # (Spine*2+30 Extra)
                                                    w_sheet.write(i, 4, v)
                                                if leaf_list1 >= 201 and leaf_list1 <= 240:
                                                    v = 12 * 2 + 30  # (Spine*2+30 Extra)
                                                    w_sheet.write(i, 4, v)
                                                if leaf_list1 >= 241 and leaf_list1 <= 280:
                                                    v = 14 * 2 + 30  # (Spine*2+30 Extra)
                                                    w_sheet.write(i, 4, v)
                                                if leaf_list1 >= 281 and leaf_list1 <= 320:
                                                    v = 16 * 2 + 30  # (Spine*2+30 Extra)
                                                    w_sheet.write(i, 4, v)
                                                if leaf_list1 >= 321 and leaf_list1 <= 360:
                                                    v = 18 * 2 + 30  # (Spine*2+30 Extra)
                                                    w_sheet.write(i, 4, v)
                                                if leaf_list1 >= 361 and leaf_list1 <= 400:
                                                    v = 20 * 2 + 30  # (Spine*2+30 Extra)
                                                    w_sheet.write(i, 4, v)
                                                if leaf_list1 >= 401 and leaf_list1 <= 440:
                                                    v = 22 * 2 + 30  # (Spine*2+30 Extra)
                                                    w_sheet.write(i, 4, v)
                                            if partnumber == '407-BBOS-US':
                                                v = 10
                                                w_sheet.write(i, 4, v)
                                            if partnumber == '210-ADUX':
                                                v = 2 + 1  # (2 Spine A and B Plus one spare)
                                                w_sheet.write(i, 4, v)
                                            if partnumber == 'A9748229':
                                                v = 2  # (For 2 Spine A and B)
                                                w_sheet.write(i, 4, v)
                                        # Spine LAYER ENTRY
                                        for i in range(2, num_rows):
                                            for j in range(0, num_cells):
                                                partnumber = sh.cell(i, j).value
                                                if partnumber == 'SPINE':
                                                    start_row_dist = i
                                                    break
                                        print("start", start_row_dist)
                                        for i in range(start_row_dist, num_rows):
                                            partnumber = sh.cell(i, 0).value
                                            if partnumber == "":
                                                end_row_dist = i
                                                break
                                        print("end", end_row_dist)
                                        print("sum", sum(leaf_list))
                                        for i in range(start_row_dist, end_row_dist + 1):
                                            print("Cumulus Large in the loop", i)
                                            partnumber = sh.cell(i, 2).value

                                            if partnumber == '407-BBOU-US':
                                                if leaf_list1 >= 1 and leaf_list1 <= 40:
                                                    v = leaf_list1 * 2 + 16  # (Leaf*2+16 Extra)
                                                    w_sheet.write(i, 4, v)
                                                if leaf_list1 >= 41 and leaf_list1 <= 80:
                                                    v = leaf_list1 * 2 + 32  # (Leaf*2+32 Extra)
                                                    w_sheet.write(i, 4, v)
                                                if leaf_list1 >= 81 and leaf_list1 <= 120:
                                                    v = leaf_list1 * 2 + 48  # (Leaf*2+48 Extra)
                                                    w_sheet.write(i, 4, v)
                                                if leaf_list1 >= 121 and leaf_list1 <= 160:
                                                    v = leaf_list1 * 2 + 64  # (Leaf*2+64 Extra)
                                                    w_sheet.write(i, 4, v)
                                                if leaf_list1 >= 161 and leaf_list1 <= 200:
                                                    v = leaf_list1 * 2 + 80  # (Leaf*2+80 Extra)
                                                    w_sheet.write(i, 4, v)
                                                if leaf_list1 >= 201 and leaf_list1 <= 240:
                                                    v = leaf_list1 * 2 + 96  # (Leaf*2+96 Extra)
                                                    w_sheet.write(i, 4, v)
                                                if leaf_list1 >= 241 and leaf_list1 <= 280:
                                                    v = leaf_list1 * 2 + 112  # (Leaf*2+112 Extra)
                                                    w_sheet.write(i, 4, v)
                                                if leaf_list1 >= 281 and leaf_list1 <= 320:
                                                    v = leaf_list1 * 2 + 128  # (Leaf*2+128 Extra)
                                                    w_sheet.write(i, 4, v)
                                                if leaf_list1 >= 321 and leaf_list1 <= 360:
                                                    v = leaf_list1 * 2 + 144  # (Leaf*2+144 Extra)
                                                    w_sheet.write(i, 4, v)
                                                if leaf_list1 >= 361 and leaf_list1 <= 400:
                                                    v = leaf_list1 * 2 + 160  # (Leaf*2+160 Extra)
                                                    w_sheet.write(i, 4, v)
                                                if leaf_list1 >= 401 and leaf_list1 <= 440:
                                                    v = leaf_list1 * 2 + 176  # (Leaf*2+176 Extra)
                                                    w_sheet.write(i, 4, v)
                                            if partnumber == '407-BBOS-US':
                                                if leaf_list1 >= 1 and leaf_list1 <= 40:
                                                    v = 4
                                                    w_sheet.write(i, 4, v)
                                                if leaf_list1 >= 41 and leaf_list1 <= 80:
                                                    v = 8
                                                    w_sheet.write(i, 4, v)
                                                if leaf_list1 >= 81 and sum(leaf_list) <= 120:
                                                    v = 12
                                                    w_sheet.write(i, 4, v)
                                                if leaf_list1 >= 121 and leaf_list1 <= 160:
                                                    v = 16
                                                    w_sheet.write(i, 4, v)
                                                if leaf_list1 >= 161 and leaf_list1 <= 200:
                                                    v = 20
                                                    w_sheet.write(i, 4, v)
                                                if leaf_list1 >= 201 and leaf_list1 <= 240:
                                                    v = 24
                                                    w_sheet.write(i, 4, v)
                                                if leaf_list1 >= 241 and leaf_list1 <= 280:
                                                    v = 28
                                                    w_sheet.write(i, 4, v)
                                                if leaf_list1 >= 281 and leaf_list1 <= 320:
                                                    v = 32
                                                    w_sheet.write(i, 4, v)
                                                if leaf_list1 >= 321 and leaf_list1 <= 360:
                                                    v = 36
                                                    w_sheet.write(i, 4, v)
                                                if leaf_list1 >= 361 and leaf_list1 <= 400:
                                                    v = 40
                                                    w_sheet.write(i, 4, v)
                                                if leaf_list1 >= 401 and leaf_list1 <= 440:
                                                    v = 44
                                                    w_sheet.write(i, 4, v)
                                            if partnumber == '210-ADUX':
                                                if leaf_list1 >= 1 and leaf_list1 <= 40:
                                                    v = 2  # (Spine A and Spine B +1)
                                                    w_sheet.write(i, 4, v)
                                                if leaf_list1 >= 41 and leaf_list1 <= 80:
                                                    v = 4  # (Spine A,b,c,d +1)
                                                    w_sheet.write(i, 4, v)
                                                if leaf_list1 >= 81 and leaf_list1 <= 120:
                                                    v = 6  # (Spine A,b,c,d,e,f +1)
                                                    w_sheet.write(i, 4, v)
                                                if leaf_list1 >= 121 and leaf_list1 <= 160:
                                                    v = 8  # (Spine A,b,c,d,e,f,g,h +1)
                                                    w_sheet.write(i, 4, v)
                                                if leaf_list1 >= 161 and leaf_list1 <= 200:
                                                    v = 10  # (Spine A,b,c,d,e,f,g,h,i,j +1)
                                                    w_sheet.write(i, 4, v)
                                                if leaf_list1 >= 201 and leaf_list1 <= 240:
                                                    v = 12  # (Spine A,b,c,d,e,f,g,h,i,j,k,l +1)
                                                    w_sheet.write(i, 4, v)
                                                if leaf_list1 >= 241 and leaf_list1 <= 280:
                                                    v = 14  # (Spine A,b,c,d,e,f,g,h,i,j,k,l,m,n +1)
                                                    w_sheet.write(i, 4, v)
                                                if leaf_list1 >= 281 and leaf_list1 <= 320:
                                                    v = 16  # (Spine A,b,c,d,e,f,g,h,i,j,k,l,m,n,o,p +1)
                                                    w_sheet.write(i, 4, v)
                                                if leaf_list1 >= 321 and leaf_list1 <= 360:
                                                    v = 18  # (Spine A,b,c,d,e,f,g,h,i,j,k,l,m,n,o,p,q,r +1)
                                                    w_sheet.write(i, 4, v)
                                                if leaf_list1 >= 361 and leaf_list1 <= 400:
                                                    v = 20  # (Spine A,b,c,d,e,f,g,h,i,j,k,l,m,n,o,p,q,r,s,t +1)
                                                    w_sheet.write(i, 4, v)
                                                if leaf_list1 >= 401 and leaf_list1 <= 440:
                                                    v = 22  # (Spine A,b,c,d,e,f,g,h,i,j,k,l,m,n,o,p,q,r,s,t,u,v +1)
                                                    w_sheet.write(i, 4, v)
                                            if partnumber == 'A9748229':
                                                if leaf_list1 >= 1 and leaf_list1 <= 40:
                                                    v = 2  # (Spine A and Spine B)
                                                    w_sheet.write(i, 4, v)
                                                if leaf_list1 >= 41 and sum(leaf_list) <= 80:
                                                    v = 4  # (Spine A,b,c,d )
                                                    w_sheet.write(i, 4, v)
                                                if leaf_list1 >= 81 and leaf_list1 <= 120:
                                                    v = 6  # (Spine A,b,c,d,e,f)
                                                    w_sheet.write(i, 4, v)
                                                if leaf_list1 >= 121 and leaf_list1 <= 160:
                                                    v = 8  # (Spine A,b,c,d,e,f,g,h)
                                                    w_sheet.write(i, 4, v)
                                                if leaf_list1 >= 161 and leaf_list1 <= 200:
                                                    v = 10  # (Spine A,b,c,d,e,f,g,h,i,j)
                                                    w_sheet.write(i, 4, v)
                                                if leaf_list1 >= 201 and leaf_list1 <= 240:
                                                    v = 12  # (Spine A,b,c,d,e,f,g,h,i,j,k,l )
                                                    w_sheet.write(i, 4, v)
                                                if leaf_list1 >= 241 and leaf_list1 <= 280:
                                                    v = 14  # (Spine A,b,c,d,e,f,g,h,i,j,k,l,m,n)
                                                    w_sheet.write(i, 4, v)
                                                if leaf_list1 >= 281 and leaf_list1 <= 320:
                                                    v = 16  # (Spine A,b,c,d,e,f,g,h,i,j,k,l,m,n,o,p)
                                                    w_sheet.write(i, 4, v)
                                                if leaf_list1 >= 321 and leaf_list1 <= 360:
                                                    v = 18  # (Spine A,b,c,d,e,f,g,h,i,j,k,l,m,n,o,p,q,r)
                                                    w_sheet.write(i, 4, v)
                                                if leaf_list1 >= 361 and leaf_list1 <= 400:
                                                    v = 20  # (Spine A,b,c,d,e,f,g,h,i,j,k,l,m,n,o,p,q,r,s,t)
                                                    w_sheet.write(i, 4, v)
                                                if leaf_list1 >= 401 and leaf_list1 <= 440:
                                                    v = 22  # (Spine A,b,c,d,e,f,g,h,i,j,k,l,m,n,o,p,q,r,s,t,u,v)
                                                    w_sheet.write(i, 4, v)

                                        # SVC SPINE LAYER ENTRY
                                        for i in range(2, num_rows):
                                            for j in range(0, num_cells):
                                                partnumber = sh.cell(i, j).value
                                                if partnumber == 'SVC SPINE':
                                                    start_row = i
                                                    break
                                        start_row1 = start_row
                                        for i in range(start_row, num_rows):
                                            partnumber = sh.cell(i, 0).value
                                            if partnumber == "":
                                                end_row = i
                                                break
                                        SVC_SPINE_end_row = end_row

                                        for i in range(start_row1, SVC_SPINE_end_row + 1):
                                            partnumber = sh.cell(i, 2).value
                                            if partnumber == '407-BBOU-US':
                                                v = 14
                                                w_sheet.write(i, 4, v)
                                            if partnumber == '407-BBOS-US':
                                                v = 80
                                                w_sheet.write(i, 4, v)
                                            if partnumber == '210-ADUX':
                                                v = 2
                                                w_sheet.write(i, 4, v)
                                            if partnumber == 'A9748229':
                                                v = 2
                                                w_sheet.write(i, 4, v)

                                        # USER ACCESS LAYER ENTRY
                                        for i in range(2, num_rows):
                                            for j in range(0, num_cells):
                                                partnumber = sh.cell(i, j).value
                                                if partnumber == 'User Access Leaves':
                                                    start_row = i
                                                    break
                                        start_row2 = start_row

                                        for j in range(start_row2, num_rows):
                                            partnumber = sh.cell(j, 0).value
                                            if partnumber == "":
                                                end_row12 = j
                                                break
                                        USER_ACCESS_Layer_end_row = end_row12

                                        for i in range(start_row2, USER_ACCESS_Layer_end_row + 1):
                                            partnumber = sh.cell(i, 2).value
                                            if partnumber == '407-BBOU-US':
                                                vv = leaf_list1 * 2
                                                w_sheet.write(i, 4, vv)

                                            if partnumber == '4610-54P-O-AC-F-US':
                                                v = leaf_list1 + 1  # (Leaf count + 1 Spare)
                                                w_sheet.write(i, 4, v)
                                            if partnumber == 'A8793201':
                                                v = leaf_list1
                                                w_sheet.write(i, 4, v)

                                        # MGMT CORE  LAYER ENTRY

                                        if flr == 1:
                                            for i in range(2, num_rows):
                                                for j in range(0, num_cells):
                                                    partnumber = sh.cell(i, j).value
                                                    if partnumber == 'Management Core Layer':
                                                        start_row = i
                                                        break
                                            start_row1 = start_row

                                            for i in range(start_row1, num_rows):
                                                partnumber = sh.cell(i, 0).value
                                                if partnumber == "":
                                                    end_row12 = i
                                                    break
                                            OOB_CORE_end_row = end_row12
                                            for i in range(start_row1, OOB_CORE_end_row):
                                                partnumber = sh.cell(i, 2).value
                                                if partnumber == '407-BBOU-US':
                                                    v = flr * 2
                                                    w_sheet.write(i, 4, v)
                                                if partnumber == '407-BBOS-US':
                                                    v = 4
                                                    w_sheet.write(i, 4, v)
                                                if partnumber == '210-ADUX':
                                                    v = 1
                                                    w_sheet.write(i, 4, v)

                                                    b = 'EDGECORE NETWORKS'
                                                    w_sheet.write(i, 1, b)
                                                    c = '4610-54P-O-AC-F-US'
                                                    w_sheet.write(i, 2, c)
                                                    d = 'AS4610-54P 48 PORT 10/100/1000BASE-T SWITCH WITH 48 POE+ PORTS,1-8 PORT SUPPORT'
                                                    w_sheet.write(i, 3, d)

                                                if partnumber == 'A9748229':
                                                    v = 1
                                                    w_sheet.write(i, 4, v)
                                                    c1 = 'A8793201'
                                                    w_sheet.write(i, 2, c1)
                                                    d1 = 'Cumulus Linux Perpetual License, 1G, 1 Year Software Updates and Support Included'
                                                    w_sheet.write(i, 3, d1)
                                            # Terminal Server Entry
                                            for i in range(2, num_rows):
                                                for j in range(0, num_cells):
                                                    partnumber = sh.cell(i, j).value
                                                    if partnumber == 'Terminal Server':
                                                        start_row = i
                                                        break
                                            start_row1 = start_row
                                            for i in range(start_row1, num_rows):
                                                partnumber = sh.cell(i, 0).value
                                                if partnumber == "":
                                                    end_row12 = i
                                                    break

                                            OOB_CORE_end_row = end_row12
                                            for i in range(start_row1, OOB_CORE_end_row):
                                                partnumber = sh.cell(i, 2).value
                                                if partnumber == 'SLC80482201S':
                                                    v = flr
                                                    w_sheet.write(i, 4, v)

                                            # MGMT Access LAYER REMOVAL

                                            for i in range(2, num_rows):
                                                for j in range(0, num_cells):
                                                    partnumber = sh.cell(i, j).value
                                                    if partnumber == 'Management Access Leaves':
                                                        start_row = i
                                                        break
                                            start_row1 = start_row

                                            for i in range(OOB_CORE_end_row, num_rows):
                                                for j in range(num_cells + 1):
                                                    w_sheet.write(i, j, "")

                                        else:
                                            for i in range(2, num_rows):
                                                for j in range(0, num_cells):
                                                    partnumber = sh.cell(i, j).value
                                                    if partnumber == 'Management Core Layer':
                                                        start_row = i
                                                        break
                                            start_row1 = start_row

                                            for i in range(start_row1, num_rows):
                                                partnumber = sh.cell(i, 0).value
                                                if partnumber == "":
                                                    end_row12 = i
                                                    break
                                            OOB_CORE_end_row = end_row12
                                            for i in range(start_row1, OOB_CORE_end_row):
                                                partnumber = sh.cell(i, 2).value
                                                if partnumber == '407-BBOU-US':
                                                    v = flr * 2 + 6
                                                    w_sheet.write(i, 4, v)
                                                if partnumber == '407-BBOS-US':
                                                    v = 4
                                                    w_sheet.write(i, 4, v)
                                                if partnumber == '210-ADUX':
                                                    v = 2
                                                    w_sheet.write(i, 4, v)

                                                if partnumber == 'A9748229':
                                                    v = 2
                                                    w_sheet.write(i, 4, v)

                                            # Terminal Server Entry
                                            for i in range(2, num_rows):
                                                for j in range(0, num_cells):
                                                    partnumber = sh.cell(i, j).value
                                                    if partnumber == 'Terminal Server':
                                                        start_row = i
                                                        break
                                            start_row1 = start_row
                                            for i in range(start_row1, num_rows):
                                                partnumber = sh.cell(i, 0).value
                                                if partnumber == "":
                                                    end_row12 = i
                                                    break
                                            OOB_CORE_end_row = end_row12
                                            for i in range(start_row1, OOB_CORE_end_row):
                                                partnumber = sh.cell(i, 2).value
                                                if partnumber == 'SLC80482201S':
                                                    v = flr + 2
                                                    w_sheet.write(i, 4, v)
                                            # MGMT Leaves LAYER ENTRY
                                            for i in range(2, num_rows):
                                                for j in range(0, num_cells):
                                                    partnumber = sh.cell(i, j).value
                                                    if partnumber == 'Management Access Leaves':
                                                        start_row = i
                                                        break
                                            start_row1 = start_row

                                            for i in range(start_row1, num_rows):
                                                partnumber = sh.cell(i, 2).value
                                                if partnumber == '407-BBOU-US':
                                                    v = flr * 2
                                                    w_sheet.write(i, 4, v)
                                                if partnumber == '4610-54P-O-AC-F-US' or partnumber == 'A8793201':
                                                    v = flr
                                                    w_sheet.write(i, 4, v)

                                        # -------------------------------------------------
                                        #       CODE FOR saving xl file after updating
                                        # -------------------------------------------------
                                        now = datetime.now()
                                        s = str('result{}.xls'.format(now.strftime("%c")))
                                        wb.save(s.replace(':', '_'))
                                        pyexcel.save_book_as(file_name=s.replace(':', '_'),
                                                             dest_file_name='result{}.xlsx'.format(
                                                                 now.strftime("%c")).replace(':',
                                                                                             '_'))
                                        s = 'result{}.xlsx'.format(now.strftime("%c")).replace(':', '_')
                                        # -------------------------------------------------
                                        #       Used openpyxl for editing the xlsx file
                                        # -------------------------------------------------
                                        import openpyxl
                                        from openpyxl.styles import Font
                                        wb = openpyxl.load_workbook(s)
                                        sheet = wb['CUMULUS Large-Site-Type']
                                        sheet1 = wb['SUMMARY-CUMULUS-BOM']
                                        # -------------------------------------------------
                                        #       CODE FOR col width(adjust the col width according to text length in it)
                                        # -------------------------------------------------
                                        dims = {}
                                        for row in sheet.rows:
                                            for cell in row:
                                                if cell.value:
                                                    dims[cell.column] = max(
                                                        (dims.get(cell.column, 0), len(str(cell.value))))
                                        for col, value in dims.items():
                                            sheet.column_dimensions[col].width = value

                                        sheet.column_dimensions['E'].width = 10

                                        sheet1.column_dimensions['A'].width = 15
                                        sheet1.column_dimensions['B'].width = 25
                                        sheet1.column_dimensions['C'].width = 25
                                        sheet1.column_dimensions['D'].width = 110
                                        sheet1.column_dimensions['E'].width = 10

                                        # -------------------------------------------------
                                        #      Cell Formatting
                                        # -------------------------------------------------

                                        top_left11_cell = sheet1['A2']

                                        top_left_cell = sheet['A2']
                                        top_left1_cell = sheet['A9']
                                        top_left2_cell = sheet['A16']
                                        top_left3_cell = sheet['A23']
                                        top_left4_cell = sheet['A29']
                                        top_left5_cell = sheet['A36']
                                        top_left6_cell = sheet['A40']
                                        sheet.merge_cells('A2:E2')
                                        sheet.merge_cells('A9:E9')
                                        sheet.merge_cells('A16:E16')
                                        sheet.merge_cells('A23:E23')
                                        sheet.merge_cells('A29:E29')
                                        sheet.merge_cells('A36:E36')
                                        sheet.merge_cells('A40:E40')

                                        sheet1.merge_cells('A2:E2')

                                        top_left_cell.alignment = Alignment(horizontal="center", vertical="center")
                                        top_left1_cell.alignment = Alignment(horizontal="center", vertical="center")
                                        top_left2_cell.alignment = Alignment(horizontal="center", vertical="center")
                                        top_left3_cell.alignment = Alignment(horizontal="center", vertical="center")
                                        top_left4_cell.alignment = Alignment(horizontal="center", vertical="center")
                                        top_left5_cell.alignment = Alignment(horizontal="center", vertical="center")
                                        top_left6_cell.alignment = Alignment(horizontal="center", vertical="center")

                                        top_left11_cell.alignment = Alignment(horizontal="center",
                                                                              vertical="center")

                                        sheet.cell(row=2, column=1).fill = PatternFill(start_color='FFFF99',
                                                                                       end_color='FFFF99',
                                                                                       fill_type='solid')
                                        sheet.cell(row=9, column=1).fill = PatternFill(start_color='FFFF99',
                                                                                       end_color='FFFF99',
                                                                                       fill_type='solid')
                                        sheet.cell(row=16, column=1).fill = PatternFill(start_color='FFFF99',
                                                                                        end_color='FFFF99',
                                                                                        fill_type='solid')
                                        sheet.cell(row=23, column=1).fill = PatternFill(start_color='FFFF99',
                                                                                        end_color='FFFF99',
                                                                                        fill_type='solid')
                                        sheet.cell(row=29, column=1).fill = PatternFill(start_color='FFFF99',
                                                                                        end_color='FFFF99',
                                                                                        fill_type='solid')
                                        sheet.cell(row=36, column=1).fill = PatternFill(start_color='FFFF99',
                                                                                        end_color='FFFF99',
                                                                                        fill_type='solid')

                                        sheet1.cell(row=2, column=1).fill = PatternFill(start_color='FFFF99',
                                                                                        end_color='FFFF99',
                                                                                        fill_type='solid')
                                        if flr == 1:
                                            pass
                                        else:
                                            sheet.cell(row=40, column=1).fill = PatternFill(start_color='FFFF99',
                                                                                            end_color='FFFF99',
                                                                                            fill_type='solid')

                                        # -------------------------------------------------
                                        #      Cell Formatting END
                                        # -------------------------------------------------

                                        # -------------------------------------------------
                                        #       Remove Not USED Sheet
                                        # -------------------------------------------------
                                        wb.remove_sheet(wb.get_sheet_by_name('DETAIL-CUMULUS-BOM-SMALL-SITE'))
                                        wb.remove_sheet(wb.get_sheet_by_name('DETAIL-CUMULUS-BOM-MEDIUM-SITE'))
                                        wb.remove_sheet(wb.get_sheet_by_name('CISCO Small-Site-Type'))
                                        wb.remove_sheet(wb.get_sheet_by_name('CISCO Medium-Site-Type'))
                                        wb.remove_sheet(wb.get_sheet_by_name('CISCO Large-Site-Type'))

                                        # -------------------------------------------------
                                        #      Removed Not USED Sheet Code END
                                        # -------------------------------------------------

                                        wb.save(s)

                                        # -------------------------------------------------
                                        #       CODE FOR MAIL STARTS
                                        # -------------------------------------------------

                                    '''subject = "An email with attachment from Python"
                                        body = "This is an email with attachment sent from Python"
                                        sender_email = "pythontesting13june@gmail.com"
                                        receiver_email = "er.akash.dhand@gmail.com"
                                        password = "python@1234"

                                        # Create a multipart message and set headers
                                        message = MIMEMultipart()
                                        message["From"] = sender_email
                                        message["To"] = receiver_email
                                        message["Subject"] = subject
                                        message["Bcc"] = receiver_email  # Recommended for mass emails

                                        # Add body to email
                                        message.attach(MIMEText(body, "plain"))

                                        filename = 'result{}.xlsx'.format(now.strftime("%c")).replace(':','_')  # In same directory as script

                                        # Open PDF file in binary mode
                                        with open(filename, "rb") as attachment:
                                            # Add file as application/octet-stream
                                            # Email client can usually download this automatically as attachment
                                            part = MIMEBase("application", "octet-stream")
                                            part.set_payload(attachment.read())

                                        # Encode file in ASCII characters to send by email
                                        encoders.encode_base64(part)

                                        # Add header as key/value pair to attachment part
                                        part.add_header(
                                            "Content-Disposition",
                                            f"attachment; filename= {filename}",
                                        )

                                        # Add attachment to message and convert message to string
                                        message.attach(part)
                                        text = message.as_string()

                                        # Log in to server using secure context and send email
                                        context = ssl.create_default_context()
                                        with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=context) as server:
                                            server.login(sender_email, password)
                                            server.sendmail(sender_email, receiver_email, text)

                                        # -------------------------------------------------
                                        #       CODE FOR MAIL ENDS
                                        # -------------------------------------------------
                                    '''
                                    # -------------------------------------------------
                                    #       XL function ends
                                    # -------------------------------------------------
                                    floor_user_count_list = []
                                    floor_name_list1 = []

                                    leaf_list = []

                                    for i in range(flr):
                                        floor_user_count_list.append(per_floor_usr_count_list[i].get())
                                        floor_name_list1.append(floor_name_list[i].get())
                                    floor_user_count_list = [int(i) for i in floor_user_count_list]

                                    for j in range(len(floor_user_count_list)):
                                        leaf_count = math.ceil(floor_user_count_list[j] / 48 + 1)
                                        leaf_list.append(leaf_count)

                                    SUM = 0
                                    for i in floor_user_count_list:
                                        SUM = SUM + i
                                    leaf_list1 = math.ceil((int(usertext) + SUM) / 48 + 1)
                                    print("Check count of leaf list", leaf_list1)

                                    txt = sc.ScrolledText(floor1_txt_frame, width=80, height=10)
                                    txt.pack()

                                    txt.insert(tk.INSERT, "Entered AMEX Site-ID :-  {}\n"
                                                          "Entered Total Number of floor :- {}\n"
                                                          "Entered Total Number of HeadCount:-{} \n"
                                                          "Entered Total Number of EUC Devices :- {}\n"
                                                          "Entered Country Name :- {}\n"
                                                          "Entered Building Name:- {}\n".format(amex_idvar, flr,
                                                                                                usertext, SUM,
                                                                                                countryVar,
                                                                                                buildind))
                                    txt.insert(tk.INSERT, "\n")

                                    txt.insert(tk.INSERT,
                                               "Total Number of Leaf required(Plus 1 spare) :- {} \n".format(
                                                   leaf_list1 + 1))

                                    print("Sum of User Count ", sum(floor_user_count_list))

                                    # -------------------------------------------------
                                    #       Call for excel function
                                    # -------------------------------------------------
                                    choice = 0
                                    # if sum(leaf_list) <= 16 and sum(leaf_list) > 1:
                                    if int(usertext) in range(1, 301):
                                        if leaf_list1 in range(0, 17):
                                            Go1_btn = tk.Button(floor2_txt_frame, text='Export',
                                                                command=XL_CUMULUS_Small_Site_Type)
                                            Go1_btn.grid(row=16, column=4)
                                        elif leaf_list1 in range(1, 41):
                                            Go1_btn = tk.Button(floor2_txt_frame, text='Export',
                                                                command=XL_CUMULUS_Medium_Site_Type)
                                            Go1_btn.grid(row=16, column=4)


                                    # elif sum(leaf_list) >= 17 and sum(leaf_list) <= 40:
                                    elif int(usertext) in range(301, 1201):
                                        if leaf_list1 in range(1, 41):
                                            Go1_btn = tk.Button(floor2_txt_frame, text='Export',
                                                                command=XL_CUMULUS_Medium_Site_Type)
                                            Go1_btn.grid(row=16, column=4)
                                        elif leaf_list1 in range(1, 841):
                                            Go1_btn = tk.Button(floor2_txt_frame, text='Export',
                                                                command=XL_CUMULUS_Large_Site_Type)
                                            Go1_btn.grid(row=16, column=4)

                                    # elif sum(leaf_list) >= 41 and sum(leaf_list) <= 440:
                                    elif int(usertext) in range(1201, 22001) and leaf_list1 in range(1, 841):
                                        Go1_btn = tk.Button(floor2_txt_frame, text='Export',
                                                            command=XL_CUMULUS_Large_Site_Type)
                                        Go1_btn.grid(row=16, column=4)

                                    Cancel_btn = tk.Button(floor2_txt_frame, text="Cancel", command=exit)
                                    Cancel_btn.grid(row=16, column=8)

                                    for i in range(len(floor_user_count_list)):
                                        D[floor_name_list1[i]] = [floor_user_count_list[i], leaf_list[i]]
                                    print(D)

                                    New_NGC_page_frame.destroy()
                                    NGC_frame_for_floor_input.destroy()

                        Go1_btn = tk.Button(NGC_frame_for_floor_input, text='Go',
                                            command=NGC_CUMULUS_Leaf_Calculation)
                        Go1_btn.grid(row=16 + i, column=4)

                        Cancel_btn = tk.Button(NGC_frame_for_floor_input, text="Cancel", command=exit)
                        Cancel_btn.grid(row=16 + i, column=5)

                    Cancel_btn = tk.Button(New_NGC_page_frame, text="BACK", command=exit)
                    Cancel_btn.grid(row=8 + 7 + 1, column=2)
                    Go1_btn = tk.Button(New_NGC_page_frame, text='NEXT225', command=NGC_CUMULUS_Function)
                    Go1_btn.grid(row=8 + 7 + 1, column=3)

                txt.insert(tk.INSERT,
                           "Total Number of Leaf required(Plus 1 spare) :- {} \n".format(leaf_list1 + 1))

                print("Sum of User Count ", sum(floor_user_count_list))

                # -------------------------------------------------
                #       Call for excel function
                # -------------------------------------------------
                choice = 0
                # if sum(leaf_list) <= 16 and sum(leaf_list) > 1:
                if int(usertext) in range(1, 301):
                    if leaf_list1 in range(0, 17):
                        Go1_btn = tk.Button(floor2_txt_frame, text='Export', command=XL_CUMULUS_Small_Site_Type)
                        Go1_btn.grid(row=16, column=4)
                    elif leaf_list1 in range(1, 41):
                        Go1_btn = tk.Button(floor2_txt_frame, text='Export', command=XL_CUMULUS_Medium_Site_Type)
                        Go1_btn.grid(row=16, column=4)


                # elif sum(leaf_list) >= 17 and sum(leaf_list) <= 40:
                elif int(usertext) in range(301, 1201):
                    if leaf_list1 in range(1, 41):
                        Go1_btn = tk.Button(floor2_txt_frame, text='Export', command=XL_CUMULUS_Medium_Site_Type)
                        Go1_btn.grid(row=16, column=4)
                    elif leaf_list1 in range(1, 841):
                        Go1_btn = tk.Button(floor2_txt_frame, text='Export', command=XL_CUMULUS_Large_Site_Type)
                        Go1_btn.grid(row=16, column=4)

                # elif sum(leaf_list) >= 41 and sum(leaf_list) <= 440:
                elif int(usertext) in range(1201, 22001) and leaf_list1 in range(1, 841):
                    Go1_btn = tk.Button(floor2_txt_frame, text='Export', command=XL_CUMULUS_Large_Site_Type)
                    Go1_btn.grid(row=16, column=4)

                Cancel_btn = tk.Button(floor2_txt_frame, text="Cancel", command=exit)
                Cancel_btn.grid(row=16, column=8)

                for i in range(len(floor_user_count_list)):
                    D[floor_name_list1[i]] = [floor_user_count_list[i], leaf_list[i]]
                print(D)

                New_NGC_page_frame.destroy()
                NGC_frame_for_floor_input.destroy()

    Go1_btn = tk.Button(NGC_frame_for_floor_input, text='Go', command=NGC_CUMULUS_Leaf_Calculation)
    Go1_btn.grid(row=16 + i, column=4)

    Cancel_btn = tk.Button(NGC_frame_for_floor_input, text="Cancel", command=exit)
    Cancel_btn.grid(row=16 + i, column=5)

def New_NGC_Cumulus_Partial_Funtion(root,font):
    New_NGC_page_frame = tk.Frame(root, bd=6, relief=tk.SUNKEN, bg='black')
    New_NGC_page_frame.place(relx=0.5, rely=0.15, relwidth=0.8, relheight=0.25, anchor='center')
    country = tk.Label(New_NGC_page_frame, text="Enter Country Name          ", font=font, fg="White", bg="Black")
    country.grid(row=0, column=1)
    country_txt = tk.Entry(New_NGC_page_frame, width=30)
    country_txt.grid(row=0, column=2)
    region = tk.Label(New_NGC_page_frame, text="Select Region         ", font=font, fg="White", bg="Black")
    region.grid(row=1, column=1)
    combo = ttk1.Combobox(New_NGC_page_frame, width=27)
    combo['values'] = ('Select ', 'USA/CAN', 'LAC', 'JAPA', 'EMEA')
    combo.current(0)
    combo.grid(row=1, column=2)
    amex_id = tk.Label(New_NGC_page_frame, text="Enter the AMEX Site-ID", font=font, fg="White", bg="Black")
    amex_id.grid(row=2, column=1)
    amex_id_txt = tk.Entry(New_NGC_page_frame, width=30)
    amex_id_txt.grid(row=2, column=2)
    building_name = tk.Label(New_NGC_page_frame, text="          Enter The Building Name", font=font, fg="White", bg="Black")
    building_name.grid(row=0, column=4)
    building_name = tk.Entry(New_NGC_page_frame, width=30)
    building_name.grid(row=0, column=5)
    usr_lable = tk.Label(New_NGC_page_frame, text="          Enter the total Headcount", font=font, fg="White", bg="Black")
    usr_lable.grid(row=1, column=4)
    usr_txt = tk.Entry(New_NGC_page_frame, width=30)
    usr_txt.grid(row=1, column=5)
    flr_lable = tk.Label(New_NGC_page_frame, text="          Total Number of IDF", font=font, fg="White", bg="Black")
    flr_lable.grid(row=2, column=4)
    flr_txt = tk.Entry(New_NGC_page_frame, width=30)
    flr_txt.grid(row=2, column=5)
    Cancel_btn = tk.Button(New_NGC_page_frame, text="BACK", command=exit)
    Cancel_btn.grid(row=8 + 7 + 1, column=2)
    Go1_btn = tk.Button(New_NGC_page_frame, text='NEXT225', command=lambda:NGC_CUMULUS_Function(root,font,country_txt,amex_id_txt,usr_txt, building_name, flr_txt, New_NGC_page_frame))
    Go1_btn.grid(row=8 + 7 + 1, column=3)



